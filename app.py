import streamlit as st
import pandas as pd
import gspread
from gspread_dataframe import get_as_dataframe, set_with_dataframe
from google.oauth2.service_account import Credentials
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, date
import time
import urllib.parse
import io
from fpdf import FPDF

# ─────────────────────────────────────────────
st.set_page_config(page_title="FinancePRO v12.2", layout="wide", page_icon="💼")

def fmt_ar(valor):
    """Formatea en pesos AR. Siempre sin decimales (los centavos no se usan)."""
    try:
        v = float(valor)
        return "$ " + f"{int(round(v)):,}".replace(",",".")
    except:
        return "$ 0"

def fmt_ar_m(valor):
    """Para métricas: sin decimales, sin abreviaciones. 
    Solo si supera 100 millones agrega x1000 como referencia."""
    try:
        v = float(valor)
        base = "$ " + f"{int(round(v)):,}".replace(",",".")
        return base
    except:
        return "$ 0"

COLORS = {
    "primary":   "#1B4F8A",
    "secondary": "#2ECC71",
    "danger":    "#E74C3C",
    "warning":   "#F39C12",
    "accent":    "#8E44AD",
    "neutral":   "#ECF0F1",
    "dark":      "#1A252F",
    "text":      "#2C3E50",
    "white":     "#FFFFFF",
    "fijo":      "#E74C3C",
    "variable":  "#F39C12",
    "chart": ["#1B4F8A","#2ECC71","#E74C3C","#F39C12",
              "#8E44AD","#1ABC9C","#E67E22","#3498DB",
              "#D35400","#27AE60","#C0392B","#2980B9"]
}

# ─────────────────────────────────────────────
# 2. CSS GLOBAL
# ─────────────────────────────────────────────
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
html, body, [class*="css"] {{ font-family: 'Inter', sans-serif !important; }}
.stApp {{ background-color: {COLORS['neutral']} !important; }}

section[data-testid="stSidebar"] {{
    background: linear-gradient(180deg, {COLORS['dark']} 0%, #243B55 100%) !important;
    min-width: 280px !important; overflow: visible !important;
    border-right: 3px solid {COLORS['primary']} !important;
}}
section[data-testid="stSidebar"] * {{ color: #FFFFFF !important; }}
section[data-testid="stSidebar"] label {{
    color: #F0F4FF !important; font-weight:700 !important;
    font-size:0.82rem !important; letter-spacing:0.06em !important;
    text-shadow: 0 1px 3px rgba(0,0,0,0.5) !important;
}}
section[data-testid="stSidebar"] .stRadio label {{
    color: #FFFFFF !important; font-size:0.95rem !important; font-weight:600 !important;
}}
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p {{ color: #CBD5E1 !important; }}
/* Selectbox y date_input del sidebar con fondo oscuro contrastado */
section[data-testid="stSidebar"] div[data-baseweb="select"] > div {{
    background-color: rgba(255,255,255,0.12) !important;
    border: 1.5px solid rgba(255,255,255,0.35) !important;
    border-radius: 8px !important;
}}
section[data-testid="stSidebar"] div[data-baseweb="select"] > div:hover {{
    background-color: rgba(255,255,255,0.2) !important;
    border-color: rgba(255,255,255,0.6) !important;
}}
section[data-testid="stSidebar"] div[data-baseweb="input"] {{
    background-color: rgba(255,255,255,0.12) !important;
    border: 1.5px solid rgba(255,255,255,0.35) !important;
    border-radius: 8px !important;
}}
/* Texto dentro de selectbox sidebar */
section[data-testid="stSidebar"] div[data-baseweb="select"] span {{ color: #FFFFFF !important; font-weight:600 !important; }}
section[data-testid="stSidebar"] input {{ color: #FFFFFF !important; font-weight:600 !important; }}

div[data-baseweb="popover"] {{ z-index:99999 !important; position:fixed !important; }}
div[data-baseweb="calendar"] {{
    z-index:99999 !important; background-color:white !important;
    border:2px solid {COLORS['primary']} !important; border-radius:12px !important;
    box-shadow:0 20px 60px rgba(0,0,0,0.4) !important; overflow:visible !important;
}}
div[data-baseweb="calendar"] * {{ color:#1A252F !important; }}
div[data-baseweb="calendar"] header {{ background-color:{COLORS['primary']} !important; border-radius:10px 10px 0 0 !important; }}
div[data-baseweb="calendar"] header button,
div[data-baseweb="calendar"] header span {{ color:white !important; font-weight:700 !important; }}
[data-baseweb="calendar"] [aria-selected="true"] div {{
    background-color:{COLORS['primary']} !important; color:white !important; border-radius:50% !important;
}}

div[data-baseweb="select"] > div, div[data-baseweb="input"], div[data-baseweb="textarea"] {{
    background-color:{COLORS['white']} !important; border:1.5px solid #BDC3C7 !important;
    border-radius:8px !important; transition:border-color 0.2s !important;
}}
div[data-baseweb="select"] > div:hover, div[data-baseweb="input"]:hover {{ border-color:{COLORS['primary']} !important; }}
input, select, textarea {{ color:{COLORS['text']} !important; }}
.stMain label, .stForm label {{ color:{COLORS['text']} !important; font-weight:600 !important; font-size:0.85rem !important; }}

[data-testid="stMetric"] {{
    background:{COLORS['white']} !important;
    padding:8px 10px !important;
    border-radius:12px !important; border:1px solid #D5DBDB !important;
    border-left:5px solid {COLORS['primary']} !important;
    box-shadow:0 2px 8px rgba(0,0,0,0.07) !important;
    transition:transform 0.2s, box-shadow 0.2s !important;
    min-width:0 !important;
}}
[data-testid="stMetric"]:hover {{ transform:translateY(-2px) !important; box-shadow:0 6px 16px rgba(0,0,0,0.12) !important; }}
[data-testid="stMetricLabel"] {{ color:#7F8C8D !important; font-size:0.68rem !important; font-weight:700 !important; letter-spacing:0.04em !important; text-transform:uppercase !important; }}
[data-testid="stMetricValue"] {{
    font-size:0.92rem !important;
    color:{COLORS['text']} !important;
    font-weight:700 !important;
    white-space:nowrap !important;
    overflow:hidden !important;
    text-overflow:ellipsis !important;
    line-height:1.3 !important;
}}

.stTabs [data-baseweb="tab-list"] {{ gap:4px !important; background-color:#D5DBDB !important; border-radius:10px !important; padding:4px !important; }}
.stTabs [data-baseweb="tab"] {{ border-radius:8px !important; font-weight:600 !important; color:#7F8C8D !important; padding:8px 16px !important; }}
.stTabs [aria-selected="true"] {{ background-color:{COLORS['primary']} !important; color:white !important; }}

.stButton button {{
    background:linear-gradient(135deg, {COLORS['primary']}, #2980B9) !important;
    color:white !important; border:none !important; border-radius:8px !important;
    font-weight:600 !important; padding:8px 20px !important;
    transition:all 0.2s !important; box-shadow:0 2px 6px rgba(27,79,138,0.3) !important;
}}
.stButton button:hover {{ transform:translateY(-1px) !important; box-shadow:0 4px 12px rgba(27,79,138,0.4) !important; }}

.streamlit-expanderHeader {{
    background-color:{COLORS['white']} !important; border:1px solid #D5DBDB !important;
    border-radius:10px !important; font-weight:600 !important; color:{COLORS['text']} !important;
}}
[data-testid="stDataFrame"] {{ border-radius:10px !important; overflow:hidden !important; }}

div[data-testid="stForm"] {{
    background:white !important; border-radius:18px !important;
    padding:32px 28px !important; box-shadow:0 24px 64px rgba(0,0,0,0.35) !important; border:none !important;
}}
div[data-testid="stForm"] label {{
    color:#1A252F !important; font-weight:700 !important;
    font-size:0.88rem !important; letter-spacing:0.03em !important;
}}
div[data-testid="stForm"] div[data-baseweb="input"] {{
    background-color: #F0F4FF !important;
    border: 2px solid #1B4F8A !important;
    border-radius: 10px !important;
}}
div[data-testid="stForm"] div[data-baseweb="input"] input {{
    color: #1A252F !important; font-weight: 600 !important; font-size: 0.95rem !important;
}}
div[data-testid="stForm"] div[data-baseweb="input"]:focus-within {{
    border-color: #8E44AD !important;
    box-shadow: 0 0 0 3px rgba(142,68,173,0.2) !important;
}}
div[data-testid="stForm"] button[kind="primaryFormSubmit"] {{
    background:linear-gradient(135deg, #1B4F8A, #2980B9) !important;
    color:white !important; border-radius:10px !important; font-size:1rem !important;
    font-weight:700 !important; padding:12px !important; border:none !important;
    box-shadow:0 4px 14px rgba(27,79,138,0.4) !important;
}}

.page-header {{
    background:linear-gradient(135deg, {COLORS['primary']} 0%, #2980B9 100%);
    color:white !important; padding:18px 24px; border-radius:14px;
    margin-bottom:20px; box-shadow:0 4px 15px rgba(27,79,138,0.25);
}}
.page-header h2 {{ margin:0; color:white !important; font-size:1.4rem !important; font-weight:700 !important; }}
.page-header span {{ font-size:0.88rem; opacity:0.85; }}

.whatsapp-btn {{
    background-color:#25D366; color:white !important; padding:7px 14px;
    border-radius:8px; text-decoration:none; font-size:0.8rem; font-weight:700;
    box-shadow:0 2px 6px rgba(37,211,102,0.35); transition:all 0.2s; display:inline-block;
}}

.rol-badge {{
    display:inline-block; background:{COLORS['secondary']}; color:white !important;
    font-size:0.7rem; font-weight:700; padding:3px 10px; border-radius:20px;
    letter-spacing:0.05em; margin-left:8px;
}}
.rol-badge.admin {{ background:{COLORS['accent']}; }}

.alerta-pendiente {{
    background:linear-gradient(135deg,#FEF9E7,#FDEBD0) !important;
    border-left:4px solid {COLORS['warning']} !important;
    border-radius:8px !important; padding:10px 14px !important; margin:6px 0 !important;
}}

.sidebar-brand {{ text-align:center; padding:16px 0 8px 0; border-bottom:1px solid rgba(255,255,255,0.1); margin-bottom:12px; }}
.sidebar-brand h1 {{ color:white !important; font-size:1.5rem !important; font-weight:800 !important; margin:0 !important; letter-spacing:-0.02em; }}
.sidebar-brand small {{ color:#95A5A6 !important; font-size:0.75rem !important; }}

.inv-card {{
    background:white; border-radius:12px; padding:16px 20px;
    border:1px solid #D5DBDB; border-left:5px solid {COLORS['accent']};
    box-shadow:0 2px 8px rgba(0,0,0,0.06); margin-bottom:10px;
}}
</style>
""", unsafe_allow_html=True)


# ── MODO OSCURO: inyección dinámica de CSS cuando está activado ──
_DARK_CSS = """
<style>
.stApp { background-color: #0F172A !important; }
[data-testid="stMain"] > div { background-color: #0F172A !important; }
[data-testid="stMetric"] { background: #1E293B !important; border-color: #334155 !important; border-left-color: #3B82F6 !important; }
[data-testid="stMetricValue"] { color: #F1F5F9 !important; }
[data-testid="stMetricLabel"] { color: #94A3B8 !important; }
.stTabs [data-baseweb="tab-list"] { background-color: #1E293B !important; }
.stTabs [data-baseweb="tab"] { color: #94A3B8 !important; }
.stTabs [aria-selected="true"] { background-color: #3B82F6 !important; color: white !important; }
.streamlit-expanderHeader { background-color: #1E293B !important; border-color: #334155 !important; color: #E2E8F0 !important; }
div[data-baseweb="select"] > div, div[data-baseweb="input"], div[data-baseweb="textarea"] { background-color: #1E293B !important; border-color: #475569 !important; }
input, select, textarea { color: #F1F5F9 !important; }
.stMain label, .stForm label { color: #CBD5E1 !important; }
[data-testid="stMarkdownContainer"] p, [data-testid="stMarkdownContainer"] span { color: #CBD5E1 !important; }
h1,h2,h3,h4 { color: #F1F5F9 !important; }
[data-testid="stDataFrame"] th { background: #0F172A !important; color: #94A3B8 !important; }
[data-testid="stDataFrame"] td { background: #1E293B !important; color: #E2E8F0 !important; }
.alerta-pendiente { background: linear-gradient(135deg,#292524,#3d2e1e) !important; }
.inv-card { background: #1E293B !important; border-color: #334155 !important; }
.page-header { box-shadow: 0 4px 20px rgba(0,0,0,0.6) !important; }
div[data-testid="stForm"] { background: #1E293B !important; box-shadow: 0 8px 32px rgba(0,0,0,0.5) !important; }
div[data-testid="stForm"] label { color: #CBD5E1 !important; }
div[data-testid="stForm"] div[data-baseweb="input"] { background-color: #0F172A !important; border-color: #3B82F6 !important; }
div[data-testid="stForm"] div[data-baseweb="input"] input { color: #F1F5F9 !important; }
.stButton button { background: linear-gradient(135deg, #1D4ED8, #2563EB) !important; }
.stRadio label { color: #E2E8F0 !important; }
/* textarea en perfil WA modo oscuro */
textarea { background-color: #1E293B !important; color: #F1F5F9 !important; border-color: #475569 !important; }
[data-testid="stNotificationContentInfo"] { background: #1E3A5F !important; color: #93C5FD !important; }
[data-testid="stNotificationContentWarning"] { background: #3D2A00 !important; color: #FCD34D !important; }
[data-testid="stNotificationContentSuccess"] { background: #052E16 !important; color: #86EFAC !important; }
[data-testid="stCaptionContainer"] { color: #64748B !important; }
[data-baseweb="popover"] [role="option"] { background: #1E293B !important; color: #E2E8F0 !important; }
[data-baseweb="popover"] [role="option"]:hover { background: #334155 !important; }
/* Todos los textos generales claros en modo oscuro */
[data-testid="stMarkdownContainer"] * { color: #E2E8F0 !important; }
[data-testid="stMarkdownContainer"] strong,
[data-testid="stMarkdownContainer"] b { color: #F8FAFC !important; }
[data-testid="stMarkdownContainer"] small { color: #94A3B8 !important; }
[data-testid="stMarkdownContainer"] code { background: #334155 !important; color: #7DD3FC !important; }
/* Alerta pendiente en dark: texto legible */
.alerta-pendiente * { color: #F1C40F !important; }
.alerta-pendiente strong { color: #FBBF24 !important; }
/* Caption / info / warning */
[data-testid="stCaption"] { color: #64748B !important; }
[data-testid="stNotificationContentInfo"] { background: #1E3A5F !important; color: #93C5FD !important; }
/* Expander content */
[data-testid="stExpanderDetails"] { background: #1E293B !important; }
[data-testid="stExpanderDetails"] * { color: #E2E8F0 !important; }
/* Divider */
hr { border-color: #334155 !important; }
/* selectbox dropdown list */
[data-baseweb="popover"] [role="option"] { background: #1E293B !important; color: #E2E8F0 !important; }
/* stMetric delta */
[data-testid="stMetricDelta"] { color: #86EFAC !important; }
/* page header text always white */
.page-header, .page-header * { color: white !important; }
/* radio buttons label dark mode */
[data-testid="stWidgetLabel"] { color: #CBD5E1 !important; }
</style>
"""

# ─────────────────────────────────────────────
# 3. DATOS
# ─────────────────────────────────────────────
URL_SHEET = "https://docs.google.com/spreadsheets/d/152P8Nuk-dlb7S_EYrodxXzAsqZNBXJ5eV2mdzXKRVn4/edit#gid=33562255"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

@st.cache_resource(show_spinner=False)
def get_gc():
    """Conexión Google autenticada — cacheada como resource (no se recrea en cada rerun)."""
    creds_dict = dict(st.secrets["connections"]["gsheets"])
    creds_dict["private_key"] = creds_dict["private_key"].replace("\\n", "\n")
    creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    return gspread.authorize(creds)

@st.cache_resource(show_spinner=False)
def get_sheet():
    """Objeto Spreadsheet cacheado — una sola apertura por sesión del servidor."""
    return get_gc().open_by_url(URL_SHEET)

def open_sheet():
    return get_sheet()

def write_ws(ws_name, df):
    """Escribe DataFrame en la hoja. Crea la hoja si no existe."""
    sh = open_sheet()
    try:
        ws = sh.worksheet(ws_name)
    except Exception:
        # La hoja no existe → crearla automáticamente
        ws = sh.add_worksheet(title=ws_name, rows=1000, cols=30)
    ws.clear()
    set_with_dataframe(ws, df, include_index=False, resize=True)
    if 'data_cache' not in st.session_state:
        st.session_state['data_cache'] = {}
    st.session_state['data_cache'][ws_name] = df.copy()

def read_ws(sh, name):
    try:
        ws = sh.worksheet(name)
    except Exception:
        return pd.DataFrame()   # hoja no existe → DataFrame vacío
    df  = get_as_dataframe(ws, evaluate_formulas=True, dtype=str)
    df  = df.dropna(how='all').reset_index(drop=True)
    df  = df.loc[:, df.columns.notna()]
    df  = df.loc[:, ~df.columns.astype(str).str.startswith('Unnamed')]
    return df

@st.cache_data(ttl=120, show_spinner="Cargando datos...")
def cargar_todo(url):
    sh = open_sheet()
    return {
        "users":  read_ws(sh, "Users"),
        "movs":   read_ws(sh, "Movimientos"),
        "config": read_ws(sh, "Config"),
        "inv":    read_ws(sh, "Inversiones"),
    }

def get_data():
    """Retorna datos: usa cache local (session_state) si existe, si no el cache de 120s."""
    base = cargar_todo(URL_SHEET)
    cache = st.session_state.get('data_cache', {})
    return {
        "users":  cache.get("Users",        base["users"]),
        "movs":   cache.get("Movimientos",  base["movs"]),
        "config": cache.get("Config",       base["config"]),
        "inv":    cache.get("Inversiones",  base["inv"]),
    }

def get_frase_semanal(df_config):
    """Lee la frase semanal desde la hoja Config."""
    try:
        if 'frase_semanal' in df_config.columns:
            _f = df_config['frase_semanal'].dropna()
            _f = _f[_f.astype(str).str.strip() != '']
            if not _f.empty:
                return str(_f.iloc[0]).strip()
    except Exception:
        pass
    return ""  

data = get_data()

if 'logged_in' not in st.session_state:
    st.session_state.update({'logged_in': False, 'user': None})

# ─────────────────────────────────────────────
# 4. LOGIN
# ─────────────────────────────────────────────
if not st.session_state['logged_in']:
    st.markdown("""
    <style>
    .stApp { background: linear-gradient(135deg, #1B4F8A 0%, #243B55 50%, #1A252F 100%) !important; }
    header[data-testid="stHeader"] { background: transparent !important; }
    </style>
    """, unsafe_allow_html=True)
    c1, c2, c3 = st.columns([1, 1.1, 1])
    with c2:
        st.markdown("""
        <div style="text-align:center; padding:48px 0 24px 0;">
            <div style="font-size:3.8rem; filter:drop-shadow(0 4px 12px rgba(0,0,0,0.3));">💼</div>
            <h1 style="color:white; font-size:2.4rem; font-weight:800; margin:10px 0 6px 0;
                       letter-spacing:-0.03em; text-shadow:0 2px 8px rgba(0,0,0,0.3);">FinancePRO</h1>
            <p style="color:#94A3B8; font-size:0.9rem; margin:0; letter-spacing:0.04em;">GESTIÓN FINANCIERA INTELIGENTE</p>
        </div>
        """, unsafe_allow_html=True)
        with st.form("login"):
            st.markdown('<p style="color:#64748B; font-size:0.8rem; font-weight:700; letter-spacing:0.1em; margin:0 0 20px 0; text-align:center;">ACCESO AL SISTEMA</p>', unsafe_allow_html=True)
            u_email = st.text_input("📧 Email", placeholder="tucuenta@email.com").lower().strip()
            u_pass  = st.text_input("🔒 Contraseña", type="password", placeholder="••••••••").strip()
            st.markdown("<br>", unsafe_allow_html=True)
            if st.form_submit_button("INGRESAR →", use_container_width=True):
                users = data['users'].copy()
                users['email']    = users['email'].astype(str).str.lower().str.strip()
                # Limpiar password: strip, quitar .0 de floats, quitar comillas extras
                users['password'] = (users['password'].astype(str)
                                     .str.strip()
                                     .str.replace(r'\.0$','',regex=True)
                                     .str.replace(r'^["\']+|["\']+$','',regex=True))
                u_pass = u_pass.strip()
                match = users[(users['email']==u_email) & (users['password']==u_pass)]
                if not match.empty:
                    u_dict = match.iloc[0].to_dict()
                    tpl_saved = str(u_dict.get('wa_template','')).strip()
                    if tpl_saved and tpl_saved not in ('nan',''):
                        st.session_state['wa_template'] = tpl_saved
                    st.session_state.update({
                        'logged_in': True,
                        'user': u_dict,
                        'login_ts': time.time()   # timestamp del login para control de 12hs
                    })
                    st.rerun()
                else:
                    st.error("❌ Email o contraseña incorrectos")
        st.markdown('<p style="text-align:center; color:#64748B; font-size:0.75rem; margin-top:24px;">v12.2 · Desarrollado con ❤️</p>', unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 5. APP PRINCIPAL
# ─────────────────────────────────────────────
else:
    # ── Verificar sesión de 12 horas ──
    _SESSION_HS = 12
    _login_ts   = st.session_state.get('login_ts', time.time())
    _elapsed_hs = (time.time() - _login_ts) / 3600
    if _elapsed_hs > _SESSION_HS:
        st.session_state.clear()
        st.cache_data.clear()
        st.session_state.pop('data_cache', None)
        st.info("⏰ Tu sesión expiró después de 12 horas. Ingresá nuevamente.")
        st.rerun()

    user        = st.session_state['user']
    es_admin    = user['rol'] == 'admin'
    df_config   = data['config'].copy()
    df_movs_raw = data['movs'].copy()
    df_inv_raw  = data['inv'].copy()

    # Columna naturaleza en Config (Fijo / Variable)
    if 'naturaleza' not in df_config.columns:
        df_config['naturaleza'] = 'Variable'

    # ── SIDEBAR ──
    with st.sidebar:
        st.markdown('<div class="sidebar-brand"><h1>💼 FinancePRO</h1><small>v12.2</small></div>', unsafe_allow_html=True)
        if es_admin:
            cl_df        = data['users'][data['users']['rol']=='cliente']
            dict_c       = pd.Series(cl_df.email.values, index=cl_df.nombre).to_dict()
            sel_nombre   = st.selectbox("👤 CLIENTE:", list(dict_c.keys()))
            cliente_mail = dict_c[sel_nombre].strip().lower()
        else:
            cliente_mail, sel_nombre = user['email'].strip().lower(), user['nombre']

        st.markdown("---")
        menu  = st.radio("📌 MENÚ", ["📊 Dashboard","💸 Movimientos","📈 Inversiones","⚙️ Perfil"])
        st.markdown("**🗓️ PERÍODO:**")
        rango = st.date_input("", [date.today().replace(day=1), date.today()], label_visibility="collapsed")
        st.markdown("---")
        rol_label = "ADMIN" if es_admin else "CLIENTE"
        st.markdown(f"""
        <div style="padding:8px 0;">
            <p style="margin:0; font-size:0.78rem; color:#95A5A6;">Usuario activo</p>
            <p style="margin:4px 0 0 0; font-weight:700; font-size:0.95rem;">{user['nombre']}</p>
            <span class="rol-badge {'admin' if es_admin else ''}">{rol_label}</span>
        </div>""", unsafe_allow_html=True)
        st.markdown("---")
        dark = st.session_state.get('dark_mode', False)
        lbl_dark = "☀️ Modo Claro" if dark else "🌙 Modo Oscuro"
        if st.button(lbl_dark, use_container_width=True):
            st.session_state['dark_mode'] = not dark
            st.rerun()
        if st.button("🚪 Cerrar Sesión", use_container_width=True):
            st.session_state.clear(); st.cache_data.clear(); st.session_state.pop('data_cache', None); st.rerun()

    # ── INYECTAR MODO OSCURO si está activo ──
    if st.session_state.get('dark_mode', False):
        st.markdown(_DARK_CSS, unsafe_allow_html=True)

    # ── PROCESAR MOVIMIENTOS ──
    df_movs_raw['monto']     = pd.to_numeric(df_movs_raw['monto'],    errors='coerce').fillna(0)
    df_movs_raw['pendiente'] = pd.to_numeric(df_movs_raw['pendiente'],errors='coerce').fillna(0)
    df_movs_raw['id']        = pd.to_numeric(df_movs_raw['id'],       errors='coerce').fillna(0).astype(int)
    # Limpiar whatsapp_contacto: quitar .0 de float, espacios, guiones
    if 'whatsapp_contacto' in df_movs_raw.columns:
        df_movs_raw['whatsapp_contacto'] = (df_movs_raw['whatsapp_contacto']
            .astype(str).str.strip()
            .str.replace(r'\.0$','',regex=True)
            .str.replace(r'[\s\-\+]','',regex=True)
            .str.replace(r'^nan$','',regex=True))

    df_c = df_movs_raw[df_movs_raw['email']==cliente_mail].copy()
    df_c['fecha_dt'] = pd.to_datetime(df_c['fecha'], dayfirst=True, errors='coerce')

    # Enriquecer con naturaleza desde Config
    # Estructura real del Sheet Config:
    #   A=categoria | B=tipo_asociado | C=subtipo (Fijo/Variable, solo gastos) | D=medios
    # El cliente carga 'categoria' en sus movimientos → mapeamos categoria → subtipo
    nat_map = {}
    if 'categoria' in df_config.columns:
        col_nat = 'subtipo' if 'subtipo' in df_config.columns else (
                  'naturaleza' if 'naturaleza' in df_config.columns else None)
        if col_nat:
            _cfg = df_config[df_config['categoria'].notna()].copy()
            _cfg['categoria'] = _cfg['categoria'].astype(str).str.strip()
            _cfg[col_nat]     = _cfg[col_nat].astype(str).str.strip()
            # Mantener primer valor por categoria para no pisar con NaN o '-'
            _cfg = _cfg[~_cfg[col_nat].isin(['', 'nan', '-', 'NaN'])]
            _cfg = _cfg.drop_duplicates(subset='categoria', keep='first')
            nat_map = dict(zip(_cfg['categoria'], _cfg[col_nat]))
    # Normalizar: solo valores válidos son 'Fijo' y 'Variable'
    def _nat(cat):
        v = nat_map.get(str(cat).strip(), 'Variable')
        return 'Fijo' if str(v).strip().lower() in ('fijo','fixed') else 'Variable'
    df_c['naturaleza'] = df_c['categoria'].apply(_nat)

    # Cargar banco de cheques del cliente (para autocompletar en gastos)
    try:
        _sh_ch = open_sheet()
        df_cheques_banco = read_ws(_sh_ch, "Cheques")
        if not df_cheques_banco.empty and 'email' in df_cheques_banco.columns:
            df_cheques_banco = df_cheques_banco[
                df_cheques_banco['email']==cliente_mail
            ].copy()
            df_cheques_banco['monto'] = pd.to_numeric(df_cheques_banco['monto'], errors='coerce').fillna(0)
        else:
            df_cheques_banco = pd.DataFrame()
    except Exception:
        df_cheques_banco = pd.DataFrame()

    # Cheques disponibles (recibidos, no usados ni vencidos)
    if not df_cheques_banco.empty and 'tipo' in df_cheques_banco.columns:
        df_ch_disponibles = df_cheques_banco[
            (df_cheques_banco['tipo']=='Cheque Recibido') &
            (df_cheques_banco.get('estado', pd.Series(['Pendiente']*len(df_cheques_banco)))
             .isin(['Pendiente','pendiente','']))
        ].copy()
    else:
        df_ch_disponibles = pd.DataFrame()

    if isinstance(rango,(list,tuple)) and len(rango)==2:
        df_f = df_c[(df_c['fecha_dt'].dt.date>=rango[0]) & (df_c['fecha_dt'].dt.date<=rango[1])].copy()
    else:
        df_f = df_c.copy()

    # Medios de pago: columna D del Config, ignorar vacíos y guiones
    medios_col = 'medios' if 'medios' in df_config.columns else None
    if medios_col:
        # Normalizar: Title Case + deduplicar (evita cheque/Cheque/CHEQUE duplicados)
        _raw_medios = [str(m).strip().title() for m in df_config[medios_col].dropna().unique().tolist()
                       if str(m).strip() not in ('', 'nan', '-')]
        # Preservar orden y deduplicar case-insensitive
        _seen = set()
        medios = []
        for _m in _raw_medios:
            if _m.lower() not in _seen:
                _seen.add(_m.lower())
                medios.append(_m)
    else:
        medios = ["Efectivo", "Transferencia", "A cuenta"]

    # Filtrar categorías del Config para este cliente (si hay columna 'email')
    if 'email' in df_config.columns:
        df_config_cliente = df_config[
            df_config['email'].astype(str).str.lower().str.strip() == cliente_mail
        ].copy()
        # Si no hay filas para este cliente, usar todas (compatibilidad hacia atrás)
        if df_config_cliente.empty:
            df_config_cliente = df_config.copy()
    else:
        df_config_cliente = df_config.copy()

    # ── HELPER: generar PDF ──
    def generar_pdf(df_periodo, df_todos, cliente_nombre, periodo_str):
        """Genera PDF con fpdf2 — compatible con servidores sin display."""

        class PDF(FPDF):
            def header(self):
                self.set_font("Helvetica", "B", 14)
                self.set_text_color(27, 79, 138)
                self.cell(0, 8, "FinancePRO", ln=False)
                self.set_font("Helvetica", "", 9)
                self.set_text_color(127, 140, 141)
                self.cell(0, 8, f"  |  {cliente_nombre}  |  {periodo_str}", ln=True)
                self.set_draw_color(27, 79, 138)
                self.set_line_width(0.5)
                self.line(10, self.get_y(), 200, self.get_y())
                self.ln(3)
            def footer(self):
                self.set_y(-12)
                self.set_font("Helvetica", "I", 7)
                self.set_text_color(149, 165, 166)
                self.cell(0, 5,
                    f"FinancePRO v12.2  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Confidencial  |  Pag. {self.page_no()}",
                    align="C")

        def set_color_primary(pdf): pdf.set_text_color(27, 79, 138)
        def set_color_dark(pdf):    pdf.set_text_color(26, 37, 47)
        def set_color_gray(pdf):    pdf.set_text_color(127, 140, 141)
        def set_color_green(pdf):   pdf.set_text_color(39, 174, 96)
        def set_color_red(pdf):     pdf.set_text_color(231, 76, 60)

        def section_title(pdf, txt):
            pdf.ln(4)
            pdf.set_fill_color(27, 79, 138)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(0, 7, f"  {txt}", ln=True, fill=True)
            pdf.ln(2)

        def tabla(pdf, headers, rows, col_w, align_last="R"):
            # Header
            pdf.set_fill_color(236, 240, 241)
            pdf.set_font("Helvetica", "B", 8)
            set_color_dark(pdf)
            for i, (h, w) in enumerate(zip(headers, col_w)):
                al = align_last if i == len(headers)-1 else "L"
                pdf.cell(w, 6, h, border=1, align=al, fill=True)
            pdf.ln()
            # Rows
            pdf.set_font("Helvetica", "", 8)
            for ri, row in enumerate(rows):
                pdf.set_fill_color(255,255,255) if ri%2==0 else pdf.set_fill_color(248,250,252)
                for i, (val, w) in enumerate(zip(row, col_w)):
                    al = align_last if i == len(row)-1 else "L"
                    pdf.cell(w, 5, str(val)[:45], border=1, align=al, fill=True)
                pdf.ln()
            pdf.ln(2)

        pdf = PDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_margins(10, 15, 10)

        # ── Métricas resumen ──
        ing   = df_periodo[df_periodo['tipo']=='Ingreso']['monto'].sum()
        gas   = df_periodo[df_periodo['tipo']=='Gasto']['monto'].sum()
        util  = ing - gas
        pend  = df_todos['pendiente'].sum()
        gas_f = df_periodo[(df_periodo['tipo']=='Gasto') & (df_periodo.get('naturaleza', pd.Series(['V']*len(df_periodo)))=='Fijo')]['monto'].sum() if 'naturaleza' in df_periodo.columns else 0
        gas_v = gas - gas_f
        margen = (util/ing*100) if ing > 0 else 0

        section_title(pdf, "Resumen del Periodo")
        metricas = [
            ("Ingresos Totales",   fmt_ar(ing),  False),
            ("Gastos Totales",     fmt_ar(gas),  False),
            ("Gastos Fijos",       fmt_ar(gas_f),False),
            ("Gastos Variables",   fmt_ar(gas_v),False),
            (f"Utilidad Neta ({margen:.1f}% margen)", fmt_ar(util), util < 0),
            ("Pendientes de Cobro",fmt_ar(pend), pend > 0),
            ("Caja Real Estimada", fmt_ar(util - pend), (util-pend) < 0),
        ]
        for label, valor, es_alerta in metricas:
            pdf.set_font("Helvetica", "B", 9)
            set_color_dark(pdf)
            pdf.cell(90, 6, f"  {label}", border="LB", fill=False)
            pdf.set_font("Helvetica", "B", 9)
            if es_alerta: set_color_red(pdf)
            else: set_color_green(pdf)
            pdf.cell(90, 6, valor, border="RB", align="R")
            pdf.ln()
        pdf.ln(3)

        # ── Ingresos por categoría ──
        df_ing = df_periodo[df_periodo['tipo']=='Ingreso']
        if not df_ing.empty:
            section_title(pdf, "Ingresos por Categoria")
            gi = df_ing.groupby('categoria')['monto'].sum().reset_index().sort_values('monto', ascending=False)
            rows_i = [[r['categoria'], fmt_ar(r['monto']), f"{r['monto']/ing*100:.1f}%"] for _,r in gi.iterrows()]
            tabla(pdf, ["Categoria","Monto","Part."], rows_i, [100, 60, 30])

        # ── Gastos por categoría ──
        df_gas = df_periodo[df_periodo['tipo']=='Gasto']
        if not df_gas.empty:
            section_title(pdf, "Gastos por Categoria")
            gg = df_gas.groupby('categoria')['monto'].sum().reset_index().sort_values('monto', ascending=False)
            rows_g = [[r['categoria'], fmt_ar(r['monto']), f"{r['monto']/gas*100:.1f}%" if gas>0 else "0%"] for _,r in gg.iterrows()]
            tabla(pdf, ["Categoria","Monto","Part."], rows_g, [100, 60, 30])

        # ── Pendientes de cobro ──
        df_pend = df_todos[df_todos['pendiente'] > 0]
        if not df_pend.empty:
            section_title(pdf, "Pendientes de Cobro")
            rows_p = [[str(r.get('nota',''))[:40], str(r.get('fecha','')), fmt_ar(r['pendiente'])]
                      for _,r in df_pend.iterrows()]
            tabla(pdf, ["Contacto","Fecha","Pendiente"], rows_p, [100, 40, 50])

        # ── Movimientos del período ──
        section_title(pdf, "Detalle de Movimientos del Periodo")
        cols_m = [c for c in ['fecha','tipo','categoria','nota','monto','medio'] if c in df_periodo.columns]
        df_show = df_periodo[cols_m].sort_values('fecha', ascending=False) if 'fecha' in cols_m else df_periodo[cols_m]
        rows_m = []
        for _,r in df_show.iterrows():
            fila = []
            for c in cols_m:
                v = r[c]
                fila.append(fmt_ar(v) if c=='monto' else str(v)[:30] if pd.notna(v) else '')
            rows_m.append(fila)
        col_w_m = [22, 18, 30, 55, 30, 25][:len(cols_m)]
        tabla(pdf, [c.capitalize() for c in cols_m], rows_m, col_w_m)

        buf = io.BytesIO(pdf.output())
        buf.seek(0)
        return buf

    # ── HELPER: exportar Excel ──    # ── HELPER: exportar Excel ──
    def generar_excel(df_periodo, df_todos, cliente_nombre, periodo_str):
        """Genera Excel multi-hoja. Intenta xlsxwriter, luego openpyxl, luego CSV zipeado."""
        buf = io.BytesIO()

        # Detectar motor disponible
        motor = None
        try:
            import xlsxwriter; motor = 'xlsxwriter'
        except ImportError:
            pass
        if not motor:
            try:
                import openpyxl; motor = 'openpyxl'
            except ImportError:
                pass

        if motor:
            with pd.ExcelWriter(buf, engine=motor) as writer:
                wb = writer.book

                # ── Helpers de formato ──
                if motor == 'xlsxwriter':
                    fmt_tit = wb.add_format({'bold':True,'font_size':14,'font_color':'#1B4F8A'})
                    fmt_hdr = wb.add_format({'bold':True,'bg_color':'#1B4F8A','font_color':'white','border':1})
                    fmt_ok  = wb.add_format({'bold':True,'font_color':'#27AE60','num_format':'$#,##0.00'})
                    fmt_bad = wb.add_format({'bold':True,'font_color':'#E74C3C','num_format':'$#,##0.00'})
                    def ws_write(ws, r, c, v, fmt=None):
                        ws.write(r, c, v, fmt)
                    def ws_setcol(ws, c1, c2, w):
                        ws.set_column(c1, c2, w)
                    def add_ws(name):
                        ws = wb.add_worksheet(name)
                        writer.sheets[name] = ws
                        return ws
                else:
                    from openpyxl.styles import Font, PatternFill, Alignment
                    fmt_tit = fmt_hdr = fmt_ok = fmt_bad = None
                    def ws_write(ws, r, c, v, fmt=None): ws.cell(row=r+1, column=c+1, value=v)
                    def ws_setcol(ws, c1, c2, w): pass
                    def add_ws(name):
                        ws = wb.create_sheet(name)
                        return ws

                # ── Hoja Resumen ──
                ws = add_ws('Resumen')
                ing   = df_periodo[df_periodo['tipo']=='Ingreso']['monto'].sum()
                gas   = df_periodo[df_periodo['tipo']=='Gasto']['monto'].sum()
                util  = ing - gas
                pend  = df_todos['pendiente'].sum()
                gas_f = df_periodo[(df_periodo['tipo']=='Gasto')&(df_periodo.get('naturaleza','Variable')=='Fijo')]['monto'].sum() if 'naturaleza' in df_periodo else 0
                gas_v = gas - gas_f
                ws_write(ws, 0, 0, f'FinancePRO — {cliente_nombre}', fmt_tit)
                ws_write(ws, 1, 0, f'Período: {periodo_str}')
                for i,(lbl,val) in enumerate([('INGRESOS',ing),('GASTOS',gas),
                    ('UTILIDAD',util),('PENDIENTE',pend),('G.FIJOS',gas_f),('G.VARIABLES',gas_v)]):
                    ws_write(ws, 3+i, 0, lbl, fmt_hdr)
                    ws_write(ws, 3+i, 1, val, fmt_ok if val>=0 else fmt_bad)
                ws_setcol(ws, 0, 0, 28); ws_setcol(ws, 1, 1, 18)

                # ── Hoja Movimientos ──
                cols_e = [c for c in ['fecha','tipo','naturaleza','categoria','nota','monto','medio','pendiente'] if c in df_periodo.columns]
                df_periodo[cols_e].to_excel(writer, sheet_name='Movimientos', index=False)

                # ── Hoja Ingresos x Categoría ──
                df_ing = df_periodo[df_periodo['tipo']=='Ingreso']
                if not df_ing.empty:
                    df_ing.groupby('categoria')['monto'].sum().reset_index()                        .to_excel(writer, sheet_name='Ing x Categoria', index=False)

                # ── Hoja Gastos Fijo/Variable ──
                df_gas = df_periodo[df_periodo['tipo']=='Gasto']
                if not df_gas.empty and 'naturaleza' in df_gas.columns:
                    df_gas.groupby(['naturaleza','categoria'])['monto'].sum().reset_index()                        .to_excel(writer, sheet_name='Gast x Naturaleza', index=False)

            buf.seek(0)
            return buf, "xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        else:
            # Fallback: CSV único con todos los datos
            cols_e = [c for c in ['fecha','tipo','naturaleza','categoria','nota','monto','medio','pendiente'] if c in df_periodo.columns]
            df_periodo[cols_e].to_csv(buf, index=False, encoding='utf-8-sig')
            buf.seek(0)
            return buf, "csv", "text/csv"


    # ══════════════════════════════════════════
    # DASHBOARD
    # ══════════════════════════════════════════
    if menu == "📊 Dashboard":
        # Métricas del PERÍODO seleccionado
        ing    = df_f[df_f['tipo']=='Ingreso']['monto'].sum()
        gas    = df_f[df_f['tipo']=='Gasto']['monto'].sum()
        gas_f  = df_f[(df_f['tipo']=='Gasto')&(df_f['naturaleza']=='Fijo')]['monto'].sum()
        gas_v  = df_f[(df_f['tipo']=='Gasto')&(df_f['naturaleza']=='Variable')]['monto'].sum()
        util   = ing - gas
        pend_t = df_c['pendiente'].sum()

        # CAJA REAL: acumulada desde siempre hasta hoy (independiente del filtro de período)
        # Opción B: cuotas/diferidos NO restan hasta que se confirma el pago
        # = Ingresos cobrados - Gastos efectivizados - Pendientes de cobro
        ing_total_acum = df_c[df_c['tipo']=='Ingreso']['monto'].sum()

        # Gastos efectivizados = gastos SIN fecha_vencimiento pendiente
        # (los que tienen fecha_vencimiento son compromisos futuros, no salen hasta confirmar)
        _df_gas = df_c[df_c['tipo']=='Gasto'].copy()
        if 'fecha_vencimiento' in _df_gas.columns:
            _fv = _df_gas['fecha_vencimiento'].astype(str).str.strip().str.upper()
            # Excluir: tienen fecha_vencimiento válida Y no están marcados como PAGADO
            _es_comprometido = _fv.notna() & (~_fv.isin(['', 'NAN', 'PAGADO']))
            gas_efectivo    = _df_gas[~_es_comprometido]['monto'].sum()
            gas_comprometido = _df_gas[_es_comprometido]['monto'].sum()
        else:
            gas_efectivo     = _df_gas['monto'].sum()
            gas_comprometido = 0.0

        caja = ing_total_acum - gas_efectivo - pend_t
        inv_m  = df_inv_raw[df_inv_raw['email']==cliente_mail]
        inv_total = pd.to_numeric(inv_m['monto'],errors='coerce').sum() if not inv_m.empty and 'monto' in inv_m.columns else 0

        per_str = f"{rango[0].strftime('%d/%m/%Y')} – {rango[1].strftime('%d/%m/%Y')}" if isinstance(rango,(list,tuple)) else ""
        st.markdown(f'<div class="page-header"><div><h2>📊 Dashboard Financiero</h2><span>{sel_nombre} · {per_str}</span></div></div>', unsafe_allow_html=True)

        # ── FRASE SEMANAL ──
        _frase = get_frase_semanal(df_config)
        if _frase and not es_admin:
            # Popup (solo 1 vez por sesión)
            _popup_key = f"frase_vista_{user['email']}"
            if not st.session_state.get(_popup_key, False):
                st.session_state[_popup_key] = True
                # Mostrar como diálogo visual con st.empty y JS
                _popup_placeholder = st.empty()
                _popup_placeholder.markdown(f"""
                <div id="frase-popup" onclick="this.style.display='none'"
                     style='position:fixed; top:0; left:0; width:100%; height:100%;
                            background:rgba(0,0,0,0.6); z-index:99999;
                            display:flex; align-items:center; justify-content:center;
                            cursor:pointer;'>
                    <div style='background:linear-gradient(135deg,#1B4F8A,#2980B9);
                                border-radius:20px; padding:40px 36px; max-width:480px;
                                text-align:center; box-shadow:0 20px 60px rgba(0,0,0,0.5);
                                animation:fadeIn 0.4s ease;'>
                        <div style='font-size:2.5rem; margin-bottom:12px;'>✨</div>
                        <p style='color:white; font-size:1.05rem; font-style:italic;
                                  line-height:1.7; margin:0; white-space:pre-line;'>{_frase}</p>
                        <p style='color:#BDC3C7; font-size:0.78rem; margin:20px 0 0 0;'>
                            Tocá en cualquier lugar para continuar</p>
                    </div>
                </div>
                <style>
                @keyframes fadeIn {{ from {{ opacity:0; transform:scale(0.95); }}
                                     to   {{ opacity:1; transform:scale(1); }} }}
                </style>
                """, unsafe_allow_html=True)

            # Banner fijo en el dashboard (siempre visible)
            _lineas_frase = _frase.split('\n')
            _texto_banner = _lineas_frase[0] if _lineas_frase else _frase
            _autor_banner = _lineas_frase[-1] if len(_lineas_frase) > 1 else ""
            st.markdown(f"""
            <div style='background:linear-gradient(135deg,#1B4F8A11,#2980B911);
                        border-left:4px solid #1B4F8A; border-radius:10px;
                        padding:10px 16px; margin-bottom:14px;
                        display:flex; align-items:center; gap:12px;'>
                <span style='font-size:1.4rem;'>✨</span>
                <div>
                    <p style='margin:0; font-style:italic; color:#1B4F8A;
                               font-size:0.88rem;'>{_texto_banner}</p>
                    {f"<p style='margin:2px 0 0 0; font-size:0.75rem; color:#7F8C8D;'>{_autor_banner}</p>" if _autor_banner else ""}
                </div>
            </div>
            """, unsafe_allow_html=True)

        # Métricas fila 1
        m1,m2,m3,m4,m5 = st.columns(5)
        m1.metric("💰 INGRESOS",    fmt_ar_m(ing))
        m2.metric("💸 GASTOS",      fmt_ar_m(gas))
        m3.metric("🔒 G. FIJOS",    fmt_ar_m(gas_f))
        m4.metric("📊 G. VARIABLES",fmt_ar_m(gas_v))
        m5.metric("📈 UTILIDAD",    fmt_ar_m(util), delta=f"{util/ing*100:.1f}% margen" if ing else None)

        # Métricas fila 2
        m6,m7,m8,_ = st.columns(4)
        m6.metric("🏦 CAJA REAL",  fmt_ar_m(caja),
                   delta=f"-{fmt_ar(pend_t)} pend." if pend_t else None,
                   help="Ingresos cobrados menos gastos efectivizados y pendientes de cobro")
        m7.metric("⏳ PENDIENTES", fmt_ar_m(pend_t),
                   help="Cobros pendientes de tus clientes")
        m8.metric("📦 INVERTIDO",  fmt_ar_m(inv_total))
        if gas_comprometido > 0:
            st.markdown(
                f"<div style='background:#FFF3CD; border-left:4px solid #F39C12; "
                f"border-radius:8px; padding:8px 14px; margin:4px 0; font-size:0.9rem;'>"
                f"⏰ <strong>Comprometido futuro:</strong> {fmt_ar(gas_comprometido)} "
                f"en cuotas/diferidos pendientes de pago — "
                f"<strong>Proyección de caja:</strong> {fmt_ar(caja - gas_comprometido)}"
                f"</div>",
                unsafe_allow_html=True
            )

        # PUNTO DE EQUILIBRIO — solo admin
        if es_admin and gas_f > 0 and ing > 0:
            mc  = (ing - gas_v) / ing if ing > 0 else 0
            pe  = gas_f / mc if mc > 0 else 0
            cob = ing / pe * 100 if pe > 0 else 0
            color_cob = '#27AE60' if cob >= 100 else '#E74C3C'
            st.markdown(f"""
            <div style="background:linear-gradient(135deg,{COLORS['primary']}12,{COLORS['accent']}10);
                        border:1.5px solid {COLORS['primary']}55; border-radius:12px;
                        padding:14px 24px; margin:14px 0; display:flex; gap:32px; flex-wrap:wrap; align-items:center;">
                <div>
                    <div style="font-size:0.72rem; font-weight:700; color:#7F8C8D; letter-spacing:0.08em; margin-bottom:4px;">🎯 PUNTO DE EQUILIBRIO <span style="background:{COLORS['accent']};color:white;font-size:0.65rem;padding:2px 7px;border-radius:8px;margin-left:6px;">SOLO ADMIN</span></div>
                    <div style="font-size:1.7rem; font-weight:800; color:{COLORS['primary']};">{fmt_ar(pe)}</div>
                </div>
                <div>
                    <div style="font-size:0.72rem; font-weight:700; color:#7F8C8D; letter-spacing:0.08em; margin-bottom:4px;">COBERTURA ACTUAL</div>
                    <div style="font-size:1.5rem; font-weight:800; color:{color_cob};">{cob:.1f}% {'✅' if cob>=100 else '⚠️'}</div>
                </div>
                <div>
                    <div style="font-size:0.72rem; font-weight:700; color:#7F8C8D; letter-spacing:0.08em; margin-bottom:4px;">MARGEN CONTRIBUCIÓN</div>
                    <div style="font-size:1.5rem; font-weight:800; color:{COLORS['accent']};">{mc*100:.1f}%</div>
                </div>
            </div>""", unsafe_allow_html=True)

        # Alerta pendientes
        deudas = df_c[df_c['pendiente']>0]
        if not deudas.empty:
            with st.expander(f"⚠️ PENDIENTES DE COBRO — Total: {fmt_ar(pend_t)}", expanded=True):
                for _,d in deudas.iterrows():
                    ca1,ca2 = st.columns([4,1])
                    ca1.markdown(f'<div class="alerta-pendiente">🔹 <strong>{d["nota"]}</strong> — Debe <strong>{fmt_ar(d["pendiente"])}</strong> <span style="color:#95A5A6;font-size:0.85rem;">({d["fecha"]})</span></div>', unsafe_allow_html=True)
                    # Template personal del usuario (clave por email) con fallback al default
                    _tpl_key = f"wa_template_{user['email']}"
                    wa_msg_tpl = st.session_state.get(_tpl_key,
                        st.session_state.get('wa_template',
                        "Hola {nombre}! Te recordamos que tenes un pago pendiente de {monto}. Muchas gracias!"))
                    msg_final = wa_msg_tpl.replace("{nombre}", str(d['nota'])).replace("{monto}", fmt_ar(d['pendiente']))
                    msg_enc   = urllib.parse.quote(msg_final)
                    # Limpiar y normalizar número argentino para wa.me
                    wa_raw = str(d['whatsapp_contacto']).strip()
                    wa_num = wa_raw.replace('+','').replace(' ','').replace('-','').replace('.','').rstrip('0123456789.')
                    wa_num = str(d['whatsapp_contacto']).strip().replace('+','').replace(' ','').replace('-','').split('.')[0]
                    # Normalizar Argentina: 549 + 10 dígitos
                    if wa_num.startswith('0'):      wa_num = '549' + wa_num[1:]
                    elif wa_num.startswith('15'):   wa_num = '549' + wa_num[2:]
                    elif wa_num.startswith('9'):    wa_num = '54'  + wa_num
                    elif not wa_num.startswith('549') and len(wa_num)==10: wa_num = '549' + wa_num
                    if wa_num and wa_num.isdigit() and len(wa_num) >= 10:
                        ca2.markdown(f'<a href="https://wa.me/{wa_num}?text={msg_enc}" target="_blank" class="whatsapp-btn">📲 Avisar</a>', unsafe_allow_html=True)
                    else:
                        ca2.warning("📵 Sin nro WA")

        # ── ALERTA PAGOS DIFERIDOS / CUOTAS PENDIENTES ──
        df_movs_raw2 = data['movs'].copy()
        df_movs_raw2['monto'] = pd.to_numeric(df_movs_raw2['monto'], errors='coerce').fillna(0)
        df_c2 = df_movs_raw2[df_movs_raw2['email']==cliente_mail].copy()
        if 'fecha_vencimiento' in df_c2.columns:
            # Excluir los ya marcados como pagados
            df_c2 = df_c2[df_c2['fecha_vencimiento'].astype(str).str.upper() != 'PAGADO']
            df_c2['fv_dt'] = pd.to_datetime(df_c2['fecha_vencimiento'], dayfirst=True, errors='coerce')
            # Para cuotas: mostrar solo la PROXIMA por grupo (nota base sin [Cuota X/N])
            if 'cuotas' in df_c2.columns:
                df_c2['_nota_base'] = df_c2['nota'].astype(str).str.replace(r' \[Cuota \d+/\d+\]','',regex=True).str.strip()
                df_c2['_es_cuota']  = pd.to_numeric(df_c2['cuotas'], errors='coerce').fillna(0) > 1
                if df_c2['_es_cuota'].any():
                    _idx_prox = (df_c2[df_c2['_es_cuota']]
                        .dropna(subset=['fv_dt'])
                        .groupby('_nota_base')['fv_dt'].idxmin().values)
                    _idx_otros = df_c2[~df_c2['_es_cuota']].index.values
                    df_c2 = df_c2.loc[list(_idx_otros) + list(_idx_prox)]
            hoy_ts = pd.Timestamp(date.today())   # mismo dtype que fv_dt -> sin TypeError
            df_pagos_venc = df_c2[
                df_c2['fv_dt'].notna() &
                (df_c2['tipo']=='Gasto') &
                (df_c2['fv_dt'] >= hoy_ts)
            ].sort_values('fv_dt')
            df_pagos_venc2 = df_c2[
                df_c2['fv_dt'].notna() &
                (df_c2['tipo']=='Gasto') &
                (df_c2['fv_dt'] < hoy_ts)
            ]
            def _lbl_cuota(pv):
                c = str(pv.get('cuotas',''))
                return f"Cuota {pv.get('cuota_num','')}/{c}" if c not in ('','nan') else 'Pago diferido'

            # cliente_tiene_cheques para el scope del dashboard
            _medios_lower_dash = [m.lower() for m in medios]
            _cliente_ch_dash   = any(x in _medios_lower_dash for x in ('cheque','e-cheq','echeq'))

            def _btn_registrar_pago(pv, key_sfx):
                """Formulario ancho inline para registrar pago de diferido/cuota."""
                _nota_pv  = str(pv.get('nota','Sin descripción') or 'Sin descripción')[:50]
                _med_def  = str(pv.get('medio','')) if str(pv.get('medio','')) in medios else medios[0]
                _mon_def  = float(pv['monto']) if pd.notna(pv['monto']) else 0.0

                with st.expander(f"✅ Confirmar pago: {_nota_pv}", expanded=False):
                    bp1, bp2, bp3, bp4 = st.columns(4)
                    med_pago = bp1.selectbox("🏦 Medio:", medios,
                        index=medios.index(_med_def) if _med_def in medios else 0,
                        key=f"med_bp_{key_sfx}")
                    fec_pago = bp2.date_input("📅 Fecha:", value=date.today(), key=f"fec_bp_{key_sfx}")
                    mon_pago = bp3.number_input("💵 Monto ($):", value=_mon_def,
                        min_value=0.0, step=1000.0, key=f"mon_bp_{key_sfx}")

                    es_cheque_pago = med_pago.lower() in ('cheque','e-cheq','echeq')
                    ch_num_bp = ch_banco_bp = ch_venc_bp = ""
                    if es_cheque_pago and _cliente_ch_dash:
                        bp4.markdown("**🏦 Cheque:**")
                        # Si el movimiento original tenía datos de cheque, prellenar
                        ch_num_bp   = bp4.text_input("N° cheque",
                            value=str(pv.get('cheque_numero','') or ''),
                            key=f"ch_num_bp_{key_sfx}")
                        ch_banco_bp = st.text_input("Banco",
                            value=str(pv.get('cheque_banco','') or ''),
                            key=f"ch_banco_bp_{key_sfx}")
                        ch_venc_date_bp = st.date_input("📅 Vencimiento cheque",
                            key=f"ch_venc_bp_{key_sfx}")
                        ch_venc_bp = ch_venc_date_bp.strftime('%d/%m/%Y')
                    else:
                        bp4.empty()

                    if st.button(f"💾 Confirmar pago — {fmt_ar(mon_pago)}",
                                 key=f"btn_bp_{key_sfx}", use_container_width=True):
                        _df_m = data['movs'].copy()
                        _df_m['monto'] = pd.to_numeric(_df_m['monto'], errors='coerce').fillna(0)
                        _id = str(pv.get('id',''))
                        if _id and 'id' in _df_m.columns:
                            _df_m.loc[_df_m['id'].astype(str)==_id, 'fecha_vencimiento'] = 'PAGADO'
                        _nuevo = pd.DataFrame([{
                            "id": int(time.time()*100),
                            "email": cliente_mail,
                            "fecha": fec_pago.strftime('%d/%m/%Y'),
                            "tipo": "Gasto",
                            "categoria": str(pv.get('categoria','') or 'Otro'),
                            "monto": mon_pago,
                            "medio": med_pago,
                            "pendiente": 0,
                            "nota": f"PAGO: {str(pv.get('nota','') or '')[:60]}",
                            "whatsapp_contacto": "",
                            "fecha_vencimiento": "",
                            "cuotas": "", "cuota_num": "",
                            "cheque_numero": ch_num_bp,
                            "cheque_banco": ch_banco_bp,
                            "cheque_venc": ch_venc_bp,
                            "usuario_log": user['nombre'],
                            "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M')
                        }])
                        write_ws("Movimientos", pd.concat([_df_m, _nuevo], ignore_index=True))
                        st.success("✅ Pago registrado correctamente")
                        st.cache_data.clear(); st.session_state.pop('data_cache', None)
                        st.rerun()

            def _render_alerta_pago(pv, key_sfx, color_izq, prefijo_html):
                """Renderiza una alerta de pago ocupando todo el ancho disponible."""
                _nota_show = str(pv.get('nota','') or 'Sin descripción')
                st.markdown(
                    f"<div style='background:#FEF9E7; border-left:4px solid {color_izq}; "
                    f"border-radius:8px; padding:10px 14px; margin:4px 0;'>"
                    f"{prefijo_html} <strong>{_nota_show}</strong> — "
                    f"{fmt_ar(pv['monto'])} — "
                    f"{'Venció' if 'Venció' in prefijo_html or 'vencido' in prefijo_html.lower() else 'Vence'}: "
                    f"{pv.get('fecha_vencimiento','')} ({_lbl_cuota(pv)})"
                    f"</div>", unsafe_allow_html=True)
                _btn_registrar_pago(pv, key_sfx)

            if not df_pagos_venc2.empty:
                with st.expander(f"🚨 PAGOS VENCIDOS SIN REALIZAR: {len(df_pagos_venc2)}", expanded=True):
                    for _i_pv, (idx_pv, pv) in enumerate(df_pagos_venc2.iterrows()):
                        _render_alerta_pago(pv, f"v_{_i_pv}", '#E74C3C',
                            "🔴")
                        st.markdown("---")

            if not df_pagos_venc.empty:
                with st.expander(f"⏰ PAGOS PRÓXIMOS A VENCER: {len(df_pagos_venc)}", expanded=True):
                    for _i_pv, (idx_pv, pv) in enumerate(df_pagos_venc.iterrows()):
                        dias_r = int((pv['fv_dt'] - hoy_ts).days)
                        color  = '#E74C3C' if dias_r <= 7 else '#F39C12'
                        _render_alerta_pago(pv, f"p_{_i_pv}", color,
                            f"<span style='color:{color};font-weight:700;'>⏰ {dias_r} días</span>")
                        st.markdown("---")

        st.divider()

        # FILTRO NATURALEZA GASTOS
        filtro_nat = st.radio("🔍 Ver gastos:", ["Todos","Solo Fijos","Solo Variables"], horizontal=True)
        df_gastos_vis = df_f[df_f['tipo']=='Gasto'].copy()
        if filtro_nat == "Solo Fijos":
            df_gastos_vis = df_gastos_vis[df_gastos_vis['naturaleza']=='Fijo']
        elif filtro_nat == "Solo Variables":
            df_gastos_vis = df_gastos_vis[df_gastos_vis['naturaleza']=='Variable']

        # FUNCIÓN DONUT INTERACTIVO
        def donut_interactivo(df_tipo, titulo, color_seq, key_sfx):
            if df_tipo.empty:
                st.info(f"Sin datos para: {titulo}"); return
            grp   = df_tipo.groupby('categoria')['monto'].sum().reset_index().sort_values('monto',ascending=False)
            total = grp['monto'].sum()
            fig = go.Figure(go.Pie(
                labels=grp['categoria'], values=grp['monto'], hole=0.52,
                marker=dict(colors=color_seq[:len(grp)], line=dict(color='white',width=2)),
                textinfo='percent+label', textfont=dict(size=11),
                hovertemplate='<b>%{label}</b><br>$%{value:,.0f}<br>%{percent}<extra></extra>',
            ))
            fig.update_layout(
                title=dict(text=f"<b>{titulo}</b>",font=dict(size=14,color=COLORS['text'])),
                showlegend=True, legend=dict(orientation="v",x=1.02,y=0.5,font=dict(size=10)),
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                margin=dict(t=50,b=20,l=20,r=120),
                annotations=[dict(text=f"<b>{fmt_ar(total)}</b>",x=0.5,y=0.5,
                                  font=dict(size=12,color=COLORS['text']),showarrow=False)]
            )
            st.plotly_chart(fig, use_container_width=True, key=f"donut_{key_sfx}")
            with st.expander(f"📋 Detalle — {titulo}"):
                grp_det = df_tipo.groupby(['categoria','medio'])['monto'].sum().reset_index()
                grp_det['%'] = (grp_det['monto']/total*100).round(1).astype(str)+'%'
                grp_det['Monto'] = grp_det['monto'].apply(fmt_ar)
                st.dataframe(grp_det[['categoria','medio','Monto','%']].rename(
                    columns={'categoria':'Categoría','medio':'Medio','%':'Part.'}
                ).sort_values('Monto',ascending=False), use_container_width=True, hide_index=True)
                fig_bar = px.bar(grp.sort_values('monto'), x='monto', y='categoria', orientation='h',
                    color='monto', color_continuous_scale=[[0,color_seq[0]],[1,color_seq[min(3,len(color_seq)-1)]]],
                    text=grp.sort_values('monto')['monto'].apply(fmt_ar),
                    labels={'monto':'Monto ($)','categoria':''})
                fig_bar.update_traces(textposition='outside',textfont_size=10)
                fig_bar.update_layout(paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',
                    coloraxis_showscale=False,showlegend=False,
                    margin=dict(t=20,b=10,l=10,r=100),height=max(220,len(grp)*42))
                st.plotly_chart(fig_bar, use_container_width=True, key=f"bar_{key_sfx}")

        cg1, cg2 = st.columns(2)
        with cg1:
            donut_interactivo(df_f[df_f['tipo']=='Ingreso'],"Ingresos por Categoría",
                ["#1B4F8A","#2980B9","#1ABC9C","#3498DB","#27AE60","#16A085"],"ing")
        with cg2:
            tit_g = f"Gastos — {filtro_nat}" if filtro_nat!="Todos" else "Gastos por Categoría"
            cols_g = ([COLORS['fijo'],"#C0392B","#E74C3C","#D35400"] if filtro_nat=="Solo Fijos"
                     else [COLORS['variable'],"#E67E22","#F39C12","#D35400"] if filtro_nat=="Solo Variables"
                     else [COLORS['danger'],COLORS['warning'],"#E67E22","#C0392B","#D35400","#F39C12"])
            donut_interactivo(df_gastos_vis, tit_g, cols_g, f"gas_{filtro_nat.replace(' ','_')}")

        # GRÁFICO FIJO vs VARIABLE
        st.divider()
        df_gas_nat = df_f[df_f['tipo']=='Gasto'].copy()
        if not df_gas_nat.empty and filtro_nat=="Todos":
            df_nat_grp = df_gas_nat.groupby('naturaleza')['monto'].sum().reset_index()
            fig_fv = go.Figure()
            for _,row in df_nat_grp.iterrows():
                clr = COLORS['fijo'] if row['naturaleza']=='Fijo' else COLORS['variable']
                fig_fv.add_trace(go.Bar(name=row['naturaleza'],x=[row['naturaleza']],y=[row['monto']],
                    marker_color=clr,width=0.4,text=fmt_ar(row['monto']),textposition='outside',
                    hovertemplate=f"<b>{row['naturaleza']}</b><br>${row['monto']:,.0f}<extra></extra>"))
            fig_fv.update_layout(
                title=dict(text="<b>Gastos Fijos vs Variables</b>",font=dict(size=14,color=COLORS['text'])),
                paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',
                showlegend=True,barmode='group',
                yaxis=dict(gridcolor='#ECF0F1',tickprefix='$'),
                margin=dict(t=50,b=30,l=20,r=20),height=280)
            cfv1,cfv2 = st.columns([1,2])
            with cfv1:
                st.plotly_chart(fig_fv,use_container_width=True,key="fv_bar")
            with cfv2:
                df_fv_d = df_gas_nat.groupby(['naturaleza','categoria'])['monto'].sum().reset_index()
                df_fv_d['Monto'] = df_fv_d['monto'].apply(fmt_ar)
                df_fv_d['%'] = (df_fv_d['monto']/df_gas_nat['monto'].sum()*100).round(1).astype(str)+'%'
                st.markdown("#### 📋 Detalle Fijo / Variable")
                st.dataframe(df_fv_d[['naturaleza','categoria','Monto','%']].rename(
                    columns={'naturaleza':'Naturaleza','categoria':'Categoría','%':'Part.'}
                ).sort_values(['Naturaleza','Monto'],ascending=[True,False]),
                use_container_width=True,hide_index=True)

        # GRÁFICO COMBINADO MENSUAL
        st.divider()
        if not df_f.empty:
            df_cp = df_f.copy()
            df_cp['Mes'] = df_cp['fecha_dt'].dt.to_period('M').dt.to_timestamp()
            df_mes_piv = df_cp.groupby(['Mes','tipo'])['monto'].sum().reset_index()\
                .pivot(index='Mes',columns='tipo',values='monto').fillna(0).reset_index()
            if not df_mes_piv.empty:
                fig_cb = go.Figure()
                if 'Ingreso' in df_mes_piv:
                    fig_cb.add_trace(go.Bar(x=df_mes_piv['Mes'],y=df_mes_piv['Ingreso'],name='Ingresos',
                        marker_color=COLORS['primary'],
                        hovertemplate='<b>Ingresos</b> %{x|%b %Y}: $%{y:,.0f}<extra></extra>'))
                if 'Gasto' in df_mes_piv:
                    fig_cb.add_trace(go.Bar(x=df_mes_piv['Mes'],y=df_mes_piv['Gasto'],name='Gastos',
                        marker_color=COLORS['danger'],
                        hovertemplate='<b>Gastos</b> %{x|%b %Y}: $%{y:,.0f}<extra></extra>'))
                if 'Ingreso' in df_mes_piv and 'Gasto' in df_mes_piv:
                    df_mes_piv['Utilidad'] = df_mes_piv['Ingreso'] - df_mes_piv['Gasto']
                    fig_cb.add_trace(go.Scatter(x=df_mes_piv['Mes'],y=df_mes_piv['Utilidad'],
                        name='Utilidad Neta',mode='lines+markers',
                        line=dict(color=COLORS['secondary'],width=3,dash='dot'),
                        marker=dict(size=8,symbol='diamond'),
                        hovertemplate='<b>Utilidad</b> %{x|%b %Y}: $%{y:,.0f}<extra></extra>'))
                fig_cb.update_layout(
                    title=dict(text="<b>Evolución Mensual — Ingresos vs Gastos</b>",font=dict(size=14,color=COLORS['text'])),
                    barmode='group',paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(dtick="M1",tickformat="%b\n%Y",gridcolor='#ECF0F1'),
                    yaxis=dict(gridcolor='#ECF0F1',tickprefix='$'),
                    legend=dict(orientation='h',y=-0.2),hovermode='x unified',
                    margin=dict(t=50,b=70),height=360)
                st.plotly_chart(fig_cb,use_container_width=True,key="combo_mensual")

        # TENDENCIA + PROYECCIÓN — solo admin
        if es_admin:
            st.divider()
            st.markdown(f"""
            <div style="background:linear-gradient(135deg,{COLORS['accent']}18,{COLORS['primary']}10);
                        border-left:4px solid {COLORS['accent']}; border-radius:10px;
                        padding:10px 16px; margin-bottom:8px;">
                <strong style="color:{COLORS['accent']};">🔐 Vista Administrador</strong>
                <span style="color:#555; font-size:0.85rem;"> — Tendencia histórica + proyección 3 meses</span>
            </div>""", unsafe_allow_html=True)

            df_evol = df_c[df_c['tipo']=='Ingreso'].copy()
            if not df_evol.empty:
                df_evol['Mes_Plot'] = df_evol['fecha_dt'].dt.to_period('M').dt.to_timestamp()
                df_eg = df_evol.groupby('Mes_Plot')['monto'].sum().reset_index().sort_values('Mes_Plot')
                fig_ev = go.Figure()
                fig_ev.add_trace(go.Scatter(
                    x=df_eg['Mes_Plot'], y=df_eg['monto'],
                    fill='tozeroy', fillcolor='rgba(142,68,173,0.1)',
                    mode='lines+markers', name='Ingresos reales',
                    line=dict(color=COLORS['accent'],width=3),
                    marker=dict(size=8,color=COLORS['accent'],line=dict(color='white',width=2)),
                    hovertemplate='<b>%{x|%b %Y}</b>: $%{y:,.0f}<extra></extra>'
                ))
                if len(df_eg) >= 3:
                    df_eg['MA3'] = df_eg['monto'].rolling(3).mean()
                    fig_ev.add_trace(go.Scatter(x=df_eg['Mes_Plot'],y=df_eg['MA3'],
                        mode='lines',name='Media móvil 3M',
                        line=dict(color=COLORS['warning'],width=2,dash='dash'),
                        hovertemplate='Media 3M: $%{y:,.0f}<extra></extra>'))
                if len(df_eg) >= 4:
                    x_n   = np.arange(len(df_eg))
                    coef  = np.polyfit(x_n, df_eg['monto'].values, 1)
                    poly  = np.poly1d(coef)
                    last  = df_eg['Mes_Plot'].iloc[-1]
                    proj_d = pd.date_range(last, periods=4, freq='MS')[1:]
                    proj_y = np.maximum(poly(np.arange(len(df_eg), len(df_eg)+3)), 0)
                    fig_ev.add_trace(go.Scatter(
                        x=[last]+list(proj_d), y=[df_eg['monto'].iloc[-1]]+list(proj_y),
                        mode='lines+markers', name='Proyección 3M',
                        line=dict(color=COLORS['secondary'],width=2.5,dash='dot'),
                        marker=dict(size=9,color=COLORS['secondary'],symbol='star',
                                    line=dict(color='white',width=1.5)),
                        hovertemplate='<b>Proyección</b> %{x|%b %Y}: $%{y:,.0f}<extra></extra>'
                    ))
                fig_ev.update_layout(
                    title=dict(text=f"<b>Tendencia + Proyección — {sel_nombre}</b>",font=dict(size=14,color=COLORS['text'])),
                    paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(dtick="M1",tickformat="%b\n%Y",gridcolor='#ECF0F1'),
                    yaxis=dict(gridcolor='#ECF0F1',tickprefix='$'),
                    hovermode='x unified',legend=dict(orientation='h',y=-0.2),
                    margin=dict(t=50,b=70),height=370)
                st.plotly_chart(fig_ev,use_container_width=True,key="tendencia_admin")
            else:
                st.info("Sin datos históricos de ingresos para este cliente.")

        # EXPORTAR REPORTES
        st.divider()
        per_str_f = f"{rango[0].strftime('%Y%m%d')}_{rango[1].strftime('%Y%m%d')}" if isinstance(rango,(list,tuple)) else "periodo"
        _col_pdf, _col_xls = st.columns(2)

        # Botón PDF
        with _col_pdf:
            try:
                _pdf_buf = generar_pdf(df_f, df_c, sel_nombre, per_str)
                st.download_button(
                    label="📄 Descargar Reporte PDF",
                    data=_pdf_buf,
                    file_name=f"FinancePRO_{sel_nombre.replace(' ','_')}_{per_str_f}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )
            except Exception as _epdf:
                st.error(f"Error generando PDF: {_epdf}")

        # Botón Excel
        with _col_xls:
            try:
                _exp_buf, _exp_ext, _exp_mime = generar_excel(df_f, df_c, sel_nombre, per_str)
                _exp_label = "📊 Descargar Reporte Excel" if _exp_ext=="xlsx" else "📊 Descargar CSV"
                st.download_button(
                    label=_exp_label,
                    data=_exp_buf,
                    file_name=f"FinancePRO_{sel_nombre.replace(' ','_')}_{per_str_f}.{_exp_ext}",
                    mime=_exp_mime,
                    use_container_width=True
                )
            except Exception as _exls:
                st.warning(f"Excel no disponible: {_exls}")

    # ══════════════════════════════════════════
    # MOVIMIENTOS
    # ══════════════════════════════════════════
    elif menu == "💸 Movimientos":
        st.markdown(f'<div class="page-header"><div><h2>💸 Movimientos</h2><span>{sel_nombre}</span></div></div>', unsafe_allow_html=True)
        # dias_alerta definida globalmente para el módulo de cheques
        dias_alerta = st.session_state.get('dias_alerta_val', 15)

        t1,t2,t3,t4 = st.tabs(["➕ Nuevo Registro","📋 Historial y Edición","💰 Gestión de Cobros","🏦 Cheques"])

        with t1:
            medios_lower = [m.lower() for m in medios]
            cliente_tiene_cheques = any(x in medios_lower for x in ('cheque','e-cheq','echeq'))

            # ── PASO 1: Tipo ──────────────────────────────────────────
            tp_v = st.radio("Tipo de movimiento:", ["Gasto","Ingreso"], horizontal=True, key="tp_carga")
            cats = (df_config_cliente[df_config_cliente['tipo_asociado']==tp_v]['categoria']
                    .dropna().unique().tolist()
                    if 'tipo_asociado' in df_config_cliente.columns else [])

            # ── PASO 2: Modalidad (solo Gasto) ───────────────────────
            forma_pago   = "Pago inmediato"
            pago_parcial = False
            if tp_v == "Gasto":
                forma_pago = st.radio("💳 Modalidad:",
                    ["Pago inmediato","Pago diferido","Pago en cuotas"],
                    horizontal=True, key="forma_pago_sel")
                pago_parcial = st.checkbox(
                    "💱 Pago parcial (parte ahora, resto diferido)", key="pago_parcial_cb")

            # ── PASO 3: Medio de pago del PRIMER cobro/pago ──────────
            st.markdown("**🏦 Medio de pago/cobro:**")
            md_v      = st.selectbox("", medios, key="medio_carga", label_visibility="collapsed")
            es_cheque = md_v.lower() in ('cheque','e-cheq','echeq') and cliente_tiene_cheques

            # Para gastos con cheque: elegir propio o de tercero
            tipo_cheque_g = "Cheque propio"
            ch_sel = None
            if tp_v == "Gasto" and es_cheque:
                tipo_cheque_g = st.radio("Tipo:",
                    ["Cheque propio","Cheque de tercero (de mi banco)"],
                    horizontal=True, key="tipo_ch_g")
                if tipo_cheque_g == "Cheque de tercero (de mi banco)":
                    if df_ch_disponibles.empty:
                        st.warning("No tenés cheques de terceros disponibles.")
                    else:
                        opciones_ch = {
                            f"N°{r['numero']} | {r.get('banco','')} | {fmt_ar(r['monto'])} | Vence: {r.get('fecha_venc','')}": r
                            for _,r in df_ch_disponibles.iterrows()
                        }
                        sel_ch_key = st.selectbox("Elegí el cheque:", list(opciones_ch.keys()), key="sel_ch_terc")
                        ch_sel = opciones_ch[sel_ch_key]
                        st.success(f"Cheque N°{ch_sel.get('numero','')} — {ch_sel.get('banco','')} — {fmt_ar(ch_sel.get('monto',0))}")

            # ── PASO 4: ¿2 medios de pago/cobro? (Ingreso Y Gasto) ──
            _lbl_2medios = "💰 Cobro con 2 medios de pago" if tp_v == "Ingreso" else "💳 Pago con 2 medios"
            st.checkbox(_lbl_2medios, key="pago_mixto_cb")
            _pago_mixto = st.session_state.get('pago_mixto_cb', False)

            # Medio del 2° cobro/pago (reactivo, fuera del form)
            _med2 = ""
            _es_cheque_m2 = False
            if _pago_mixto:
                _lbl_med2 = "🏦 2° Medio de cobro:" if tp_v == "Ingreso" else "🏦 2° Medio de pago:"
                _medio2_lista = [m for m in medios if m != md_v]
                _med2 = st.selectbox(_lbl_med2, _medio2_lista, key="medio2_outer")
                _es_cheque_m2 = _med2.lower() in ('cheque','e-cheq','echeq') and cliente_tiene_cheques
                # Para Gasto con 2° medio cheque: tipo de cheque del 2° pago
                _tipo_ch2 = "Cheque propio"
                _ch2_sel  = None
                if _es_cheque_m2 and tp_v == "Gasto":
                    _tipo_ch2 = st.radio("Tipo de cheque (2° pago):",
                        ["Cheque propio","Cheque de tercero (de mi banco)"],
                        horizontal=True, key="tipo_ch2_g")
                    if _tipo_ch2 == "Cheque de tercero (de mi banco)":
                        if df_ch_disponibles.empty:
                            st.warning("No tenés cheques de terceros disponibles.")
                        else:
                            _ops_ch2 = {
                                f"N°{r['numero']} | {r.get('banco','')} | {fmt_ar(r['monto'])}": r
                                for _,r in df_ch_disponibles.iterrows()
                            }
                            _sel_ch2_key = st.selectbox("Elegí el cheque (2° pago):", list(_ops_ch2.keys()), key="sel_ch2_terc")
                            _ch2_sel = _ops_ch2[_sel_ch2_key]
                            st.success(f"Cheque N°{_ch2_sel.get('numero','')} — {fmt_ar(_ch2_sel.get('monto',0))}")

            st.markdown("---")

            # ══════════════════════════════════════════════════════════
            # FORM — datos del movimiento
            # ══════════════════════════════════════════════════════════
            with st.form("f_carga", clear_on_submit=True):

                # ── BLOQUE 1: datos del primer cobro/pago ────────────
                if _pago_mixto:
                    _lbl_b1 = f"#### 📋 1° Cobro — {md_v}" if tp_v == "Ingreso" else f"#### 📋 1° Pago — {md_v}"
                    st.markdown(_lbl_b1)
                else:
                    st.markdown("#### 📋 Datos del movimiento")

                b1c1, b1c2, b1c3 = st.columns(3)
                f_v   = b1c1.date_input("📅 Fecha")
                cat_v = b1c2.selectbox("🏷️ Categoría", cats)
                mon_v = b1c3.number_input("💵 Monto ($)", min_value=0.0, step=1000.0)
                nt_v  = st.text_input("📝 Nota / Nombre")

                # Pendiente y WA (Ingreso)
                pn_v = wa_v = 0.0
                if tp_v == "Ingreso":
                    ip1, ip2 = st.columns(2)
                    pn_v = ip1.number_input("⏳ Pendiente de cobro ($)", min_value=0.0, step=100.0)
                    wa_v = ip2.text_input("📱 WhatsApp del deudor (549...)")

                # Cheque del 1° cobro (Ingreso)
                ch_num_v = ch_banco_v = ch_librador_v = ch_venc_v = ""
                if tp_v == "Ingreso" and es_cheque:
                    st.markdown("**🏦 Datos del cheque recibido:**")
                    ck1,ck2,ck3,ck4 = st.columns(4)
                    ch_num_v      = ck1.text_input("N° cheque")
                    ch_banco_v    = ck2.text_input("Banco emisor")
                    ch_librador_v = ck3.text_input("Librador")
                    ch_venc_date  = ck4.date_input("📅 Vencimiento")
                    ch_venc_v     = ch_venc_date.strftime('%d/%m/%Y')

                # Cheque del 1° pago (Gasto — propio)
                if tp_v == "Gasto" and es_cheque and tipo_cheque_g == "Cheque propio":
                    st.markdown("**🏦 Datos del cheque:**")
                    gk1,gk2,gk3 = st.columns(3)
                    ch_num_v      = gk1.text_input("N° cheque")
                    ch_banco_v    = gk2.text_input("Mi banco")
                    ch_venc_gp    = gk3.date_input("📅 Fecha cheque")
                    ch_venc_v     = ch_venc_gp.strftime('%d/%m/%Y')
                    ch_librador_v = user['nombre']
                elif tp_v == "Gasto" and es_cheque and ch_sel is not None:
                    ch_num_v      = str(ch_sel.get('numero',''))
                    ch_banco_v    = str(ch_sel.get('banco',''))
                    ch_librador_v = str(ch_sel.get('librador',''))
                    ch_venc_v     = str(ch_sel.get('fecha_venc',''))

                # Modalidades de pago diferido / cuotas / parcial (Gasto)
                fecha_venc_v    = ""
                cuotas_v        = 1
                cuota_num_v     = 1
                monto_parcial_v = 0.0
                fecha_parcial_v = ""

                if tp_v == "Gasto":
                    if forma_pago == "Pago diferido":
                        fd1,fd2 = st.columns(2)
                        fecha_venc_v = fd1.date_input("📅 Vencimiento del pago").strftime('%d/%m/%Y')
                        fd2.warning("Se generará una alerta antes de esta fecha.")
                    elif forma_pago == "Pago en cuotas":
                        cq1,cq2,cq3 = st.columns(3)
                        cuotas_v    = int(cq1.number_input("N° total de cuotas", min_value=2, max_value=60, value=3, step=1))
                        cuota_num_v = int(cq2.number_input("N° cuota a registrar", min_value=1, max_value=cuotas_v, value=1, step=1))
                        fecha_venc_v = cq3.date_input("📅 Vencimiento 1° cuota").strftime('%d/%m/%Y')
                        st.info(f"Cuota {cuota_num_v}/{cuotas_v} — Valor estimado por cuota: {fmt_ar(mon_v/cuotas_v if cuotas_v else mon_v)}")
                    if st.session_state.get('pago_parcial_cb', False):
                        st.markdown("**💱 Detalle del pago parcial:**")
                        pp1,pp2,pp3 = st.columns(3)
                        monto_parcial_v = pp1.number_input("Monto que pagás ahora ($)",
                            min_value=0.0,
                            max_value=float(mon_v) if mon_v > 0 else 999999999.0,
                            step=1000.0, key="monto_parcial_field")
                        fecha_parcial_v = pp2.date_input("📅 Vencimiento del resto",
                            key="fecha_parcial_field").strftime('%d/%m/%Y')
                        pp3.metric("Resto diferido", fmt_ar(max(0, mon_v - monto_parcial_v)))

                # ── BLOQUE 2: 2° cobro/pago (si pago mixto activo) ──
                mon2_v      = 0.0
                ch2_num_v   = ch2_banco_v = ch2_lib_v = ch2_venc_v = ""

                if _pago_mixto:
                    st.markdown("---")
                    _lbl_b2 = f"#### 📋 2° Cobro — {_med2}" if tp_v == "Ingreso" else f"#### 📋 2° Pago — {_med2}"
                    st.markdown(_lbl_b2)
                    _lbl_mon2 = "💵 Monto 2° cobro ($)" if tp_v == "Ingreso" else "💵 Monto 2° pago ($)"
                    mon2_v = st.number_input(_lbl_mon2, min_value=0.0, step=1000.0, key="mon2_field")
                    if _es_cheque_m2:
                        _lbl_ch2 = "**🏦 Datos del cheque recibido (2° cobro):**" if tp_v == "Ingreso" else "**🏦 Datos del cheque (2° pago):**"
                        st.markdown(_lbl_ch2)
                        ch2k1,ch2k2,ch2k3,ch2k4 = st.columns(4)
                        # Si es cheque de tercero en Gasto y ya se seleccionó, prellenar
                        _ch2_num_pre  = str(_ch2_sel.get('numero','')) if _ch2_sel else ""
                        _ch2_banco_pre= str(_ch2_sel.get('banco',''))  if _ch2_sel else ""
                        _ch2_lib_pre  = str(_ch2_sel.get('librador','')) if _ch2_sel else ""
                        ch2_num_v   = ch2k1.text_input("N° cheque (2°)",  value=_ch2_num_pre,  key="ch2_num_f")
                        ch2_banco_v = ch2k2.text_input("Banco (2°)",      value=_ch2_banco_pre, key="ch2_banco_f")
                        ch2_lib_v   = ch2k3.text_input("Librador (2°)",   value=_ch2_lib_pre,  key="ch2_lib_f")
                        ch2_venc_v  = ch2k4.date_input("📅 Vencimiento (2°)", key="ch2_venc_f").strftime('%d/%m/%Y')

                # ── GUARDAR ──────────────────────────────────────────
                st.markdown("<br>", unsafe_allow_html=True)
                if st.form_submit_button("💾 GUARDAR REGISTRO", use_container_width=True):
                    dup = df_c[
                        (df_c['fecha']==f_v.strftime('%d/%m/%Y')) &
                        (df_c['categoria']==cat_v) &
                        (df_c['monto']==mon_v) &
                        (df_c['nota']==nt_v)
                    ]
                    if not dup.empty:
                        st.warning("⚠️ Ya existe un registro similar. Verificá antes de guardar.")
                    else:
                        # Armar nota enriquecida
                        nota_extra = nt_v
                        if es_cheque and ch_num_v:
                            nota_extra = f"{nt_v} [Ch N°{ch_num_v} {ch_banco_v} vto:{ch_venc_v}]".strip()
                        if tp_v=="Gasto" and cuotas_v > 1:
                            nota_extra = f"{nota_extra} [Cuota {cuota_num_v}/{cuotas_v}]"
                        if pago_parcial and monto_parcial_v > 0:
                            nota_extra = f"{nota_extra} [Parcial: {fmt_ar(monto_parcial_v)} ahora]"

                        _ts_now  = datetime.now().strftime('%Y-%m-%d %H:%M')
                        _base_id = int(time.time()*100)

                        # Generar cuotas automáticamente
                        if tp_v == "Gasto" and cuotas_v > 1:
                            from dateutil.relativedelta import relativedelta
                            _fv_base = datetime.strptime(fecha_venc_v, '%d/%m/%Y') if fecha_venc_v else datetime.now()
                            _monto_cuota = round(mon_v / cuotas_v, 2)
                            nuevas_cuotas = []
                            for _nc in range(cuotas_v):
                                _fv_c = (_fv_base + relativedelta(months=_nc)).strftime('%d/%m/%Y')
                                nuevas_cuotas.append({
                                    "id": _base_id + _nc, "email": cliente_mail,
                                    "fecha": f_v.strftime('%d/%m/%Y'), "tipo": tp_v,
                                    "categoria": cat_v, "monto": _monto_cuota,
                                    "medio": md_v, "pendiente": 0,
                                    "nota": f"{nt_v} [Cuota {_nc+1}/{cuotas_v}]",
                                    "whatsapp_contacto": "",
                                    "fecha_vencimiento": _fv_c,
                                    "cuotas": cuotas_v, "cuota_num": _nc+1,
                                    "cheque_numero":"","cheque_banco":"","cheque_venc":"",
                                    "usuario_log": user['nombre'], "timestamp": _ts_now
                                })
                            nueva = pd.DataFrame(nuevas_cuotas)
                        else:
                            nueva = pd.DataFrame([{
                                "id": _base_id, "email": cliente_mail,
                                "fecha": f_v.strftime('%d/%m/%Y'), "tipo": tp_v,
                                "categoria": cat_v, "monto": mon_v, "medio": md_v,
                                "pendiente": pn_v, "nota": nota_extra,
                                "whatsapp_contacto": wa_v,
                                "fecha_vencimiento": fecha_venc_v,
                                "cuotas": cuotas_v if cuotas_v > 1 else "",
                                "cuota_num": cuota_num_v if cuotas_v > 1 else "",
                                "cheque_numero": ch_num_v, "cheque_banco": ch_banco_v,
                                "cheque_venc": ch_venc_v,
                                "usuario_log": user['nombre'], "timestamp": _ts_now
                            }])

                        df_concat = pd.concat([df_movs_raw, nueva], ignore_index=True)

                        # 2° cobro/pago (pago mixto)
                        if _pago_mixto and _med2 and mon2_v > 0:
                            _sufijo2 = "2° cobro" if tp_v == "Ingreso" else "2° pago"
                            _nota2 = f"{nt_v} [{_sufijo2}: {_med2}]"
                            if ch2_num_v:
                                _nota2 = f"{_nota2} [Ch N°{ch2_num_v} {ch2_banco_v} vto:{ch2_venc_v}]"
                            nueva2 = pd.DataFrame([{
                                "id": _base_id+50, "email": cliente_mail,
                                "fecha": f_v.strftime('%d/%m/%Y'), "tipo": tp_v,
                                "categoria": cat_v, "monto": mon2_v, "medio": _med2,
                                "pendiente": 0, "nota": _nota2,
                                "whatsapp_contacto": wa_v if tp_v == "Ingreso" else "",
                                "fecha_vencimiento": "",
                                "cuotas":"","cuota_num":"",
                                "cheque_numero": ch2_num_v, "cheque_banco": ch2_banco_v,
                                "cheque_venc": ch2_venc_v,
                                "usuario_log": user['nombre'], "timestamp": _ts_now
                            }])
                            df_concat = pd.concat([df_concat, nueva2], ignore_index=True)

                        write_ws("Movimientos", df_concat)

                        # Pago parcial → segundo registro diferido
                        if pago_parcial and monto_parcial_v > 0 and tp_v == "Gasto":
                            resto = mon_v - monto_parcial_v
                            if resto > 0:
                                nueva_resto = pd.DataFrame([{
                                    "id": _base_id+1, "email": cliente_mail,
                                    "fecha": f_v.strftime('%d/%m/%Y'), "tipo": "Gasto",
                                    "categoria": cat_v, "monto": resto, "medio": md_v,
                                    "pendiente": 0,
                                    "nota": f"{nt_v} [Resto diferido de pago parcial]",
                                    "whatsapp_contacto": "",
                                    "fecha_vencimiento": fecha_parcial_v,
                                    "cuotas":"","cuota_num":"",
                                    "cheque_numero":"","cheque_banco":"","cheque_venc":"",
                                    "usuario_log": user['nombre'], "timestamp": _ts_now
                                }])
                                write_ws("Movimientos",
                                    pd.concat([df_concat, nueva_resto], ignore_index=True))

                        # Cheque recibido (1° cobro) → banco de cheques
                        if tp_v == "Ingreso" and es_cheque and ch_num_v:
                            try:
                                _df_ch = read_ws(open_sheet(), "Cheques")
                            except Exception:
                                _df_ch = pd.DataFrame()
                            _nch = pd.DataFrame([{
                                "id": _base_id+2, "email": cliente_mail,
                                "tipo":"Cheque Recibido","numero":ch_num_v,
                                "banco":ch_banco_v,"librador":ch_librador_v,
                                "monto":mon_v,
                                "fecha_venc":ch_venc_v if ch_venc_v else f_v.strftime('%d/%m/%Y'),
                                "nota":nt_v,"estado":"Pendiente",
                                "usuario_log":user['nombre'],"timestamp":_ts_now
                            }])
                            write_ws("Cheques", pd.concat([_df_ch, _nch], ignore_index=True))

                        # Cheque recibido (2° cobro) → banco de cheques
                        if _pago_mixto and _es_cheque_m2 and ch2_num_v:
                            try:
                                _df_ch2 = read_ws(open_sheet(), "Cheques")
                            except Exception:
                                _df_ch2 = pd.DataFrame()
                            _nch2 = pd.DataFrame([{
                                "id": _base_id+3, "email": cliente_mail,
                                "tipo":"Cheque Recibido","numero":ch2_num_v,
                                "banco":ch2_banco_v,"librador":ch2_lib_v,
                                "monto":mon2_v,
                                "fecha_venc":ch2_venc_v if ch2_venc_v else f_v.strftime('%d/%m/%Y'),
                                "nota":f"{nt_v} [2do medio]","estado":"Pendiente",
                                "usuario_log":user['nombre'],"timestamp":_ts_now
                            }])
                            write_ws("Cheques", pd.concat([_df_ch2, _nch2], ignore_index=True))

                        # Cheque de tercero usado → marcar como Usado
                        if tp_v == "Gasto" and es_cheque and ch_sel is not None:
                            try:
                                _df_chu = read_ws(open_sheet(), "Cheques")
                                _mask = (_df_chu['numero'].astype(str)==str(ch_num_v)) &                                         (_df_chu['email']==cliente_mail)
                                _df_chu.loc[_mask,'estado'] = 'Usado en pago'
                                write_ws("Cheques", _df_chu)
                            except Exception:
                                pass

                        st.cache_data.clear()
                        st.session_state.pop('data_cache', None)
                        st.success("✅ Registro guardado correctamente")
                        st.rerun()
        with t2:
            # ── Filtros de historial ──
            fh1, fh2, fh3 = st.columns(3)
            tipos_disp  = ["Todos"] + sorted(df_f['tipo'].dropna().unique().tolist())
            medios_disp = ["Todos"] + sorted(df_f['medio'].dropna().unique().tolist())
            nats_disp   = ["Todos", "Fijo", "Variable"]

            filtro_tipo_h  = fh1.selectbox("🔍 Tipo",        tipos_disp,  key="fh_tipo")
            filtro_medio_h = fh2.selectbox("💳 Medio de pago", medios_disp, key="fh_medio")
            filtro_nat_h   = fh3.selectbox("📌 Naturaleza",   nats_disp,   key="fh_nat")

            df_hist = df_f.copy()
            if filtro_tipo_h  != "Todos": df_hist = df_hist[df_hist['tipo']    == filtro_tipo_h]
            if filtro_medio_h != "Todos": df_hist = df_hist[df_hist['medio']   == filtro_medio_h]
            if filtro_nat_h   != "Todos": df_hist = df_hist[df_hist['naturaleza'] == filtro_nat_h]

            st.caption(f"Mostrando **{len(df_hist)}** registros de {len(df_f)} totales en el período")

            df_show = df_hist.sort_values('fecha_dt',ascending=False)[['fecha','tipo','naturaleza','nota','categoria','monto','medio']].copy()
            df_show['monto'] = df_show['monto'].apply(fmt_ar)
            st.dataframe(df_show.rename(columns={'fecha':'Fecha','tipo':'Tipo','naturaleza':'Naturaleza',
                'nota':'Descripción','categoria':'Categoría','monto':'Monto','medio':'Medio'}),
                use_container_width=True, hide_index=True)
            st.markdown("---")
            st.markdown("#### ✏️ Editar / Eliminar")
            op_edit = {f"📝 {r['nota']} | 📅 {r['fecha']} | {fmt_ar(r['monto'])}": r['id']
                       for _,r in df_f.sort_values('fecha_dt',ascending=False).iterrows()}
            sel_ed = st.selectbox("🔍 Registro:", ["--- Seleccionar ---"]+list(op_edit.keys()))
            if sel_ed != "--- Seleccionar ---":
                id_t = op_edit[sel_ed]
                curr = df_movs_raw[df_movs_raw['id']==id_t].iloc[0]
                with st.form("f_edit"):
                    ce1,ce2,ce3 = st.columns(3)
                    ef  = ce1.date_input("📅 Fecha",datetime.strptime(curr['fecha'],'%d/%m/%Y'))
                    et  = ce2.selectbox("Tipo",["Ingreso","Gasto"],index=0 if curr['tipo']=="Ingreso" else 1)
                    cats_e = df_config_cliente[df_config_cliente['tipo_asociado']==et]['categoria'].dropna().unique().tolist() if 'tipo_asociado' in df_config_cliente.columns else []
                    ec  = ce3.selectbox("Categoría",cats_e)
                    em  = st.number_input("💵 Monto",value=float(curr['monto']))
                    en  = st.text_input("📝 Nota",value=curr['nota'])
                    ep  = st.number_input("⏳ Pendiente",value=float(curr['pendiente']))
                    # Campo WhatsApp editable
                    wa_curr = str(curr.get('whatsapp_contacto','')) if 'whatsapp_contacto' in curr.index else ''
                    ewa = st.text_input("📱 WhatsApp (549...)", value=wa_curr,
                                        help="Solo digitos: 549XXXXXXXXXX")
                    bu,bd = st.columns(2)
                    if bu.form_submit_button("💾 ACTUALIZAR",use_container_width=True):
                        upd_cols = ['fecha','tipo','categoria','monto','pendiente','nota','whatsapp_contacto']
                        upd_vals = [ef.strftime('%d/%m/%Y'),et,ec,em,ep,en,ewa]
                        for col,val in zip(upd_cols, upd_vals):
                            if col in df_movs_raw.columns:
                                df_movs_raw.loc[df_movs_raw['id']==id_t, col] = val
                        write_ws("Movimientos", df_movs_raw)
                        st.cache_data.clear(); st.session_state.pop('data_cache', None); st.success("✅ Actualizado"); st.rerun()
                    if bd.form_submit_button("🗑️ ELIMINAR",use_container_width=True):
                        df_movs_raw = df_movs_raw[df_movs_raw['id']!=id_t]
                        write_ws("Movimientos", df_movs_raw)
                        st.cache_data.clear(); st.session_state.pop('data_cache', None); st.rerun()

        with t3:
            deudas_v = df_c[df_c['pendiente']>0]
            if deudas_v.empty:
                st.success("✅ Sin cobros pendientes.")
            else:
                st.markdown(f"#### 💰 Total pendiente: {fmt_ar(deudas_v['pendiente'].sum())}")
                for _,r in deudas_v.iterrows():
                    with st.expander(f"💰 {r['nota']} — Debe {fmt_ar(r['pendiente'])} ({r['fecha']})"):
                        cp1,cp2 = st.columns(2)
                        m_c  = cp1.number_input("Cobrar hoy:",max_value=float(r['pendiente']),key=f"c_{r['id']}",step=100.0)
                        md_c = cp1.selectbox("Medio:",medios,key=f"md_{r['id']}")
                        if cp2.button("✅ Registrar",key=f"b_{r['id']}"):
                            df_movs_raw.loc[df_movs_raw['id']==r['id'],'pendiente'] -= m_c
                            pago = pd.DataFrame([{"id":int(time.time()*100),"email":cliente_mail,
                                "fecha":date.today().strftime('%d/%m/%Y'),"tipo":"Ingreso",
                                "categoria":"Cobro Deuda","monto":m_c,"medio":md_c,
                                "pendiente":0,"nota":f"Cobro a: {r['nota']}","usuario_log":user['nombre']}])
                            write_ws("Movimientos", pd.concat([df_movs_raw,pago],ignore_index=True))
                            st.cache_data.clear(); st.session_state.pop('data_cache', None); st.rerun()

        # ═════════════════ TAB CHEQUES ═════════════════
        with t4:
            st.markdown("#### 🏦 Banco de Cheques")
            st.info("Los cheques se registran automáticamente al cargar un movimiento con medio **Cheque** o **E-Cheq**.")

            # Días de alerta configurable (actualiza session_state para persistir)
            dias_alerta = st.number_input(
                "⚙️ Alertar cuando falten X días para vencimiento del cheque:",
                min_value=1, max_value=90,
                value=st.session_state.get('dias_alerta_val', 15), step=1,
                key="dias_alerta_input"
            )
            st.session_state['dias_alerta_val'] = int(dias_alerta)

            # ── Cargar cheques existentes ──
            try:
                df_ch = read_ws(open_sheet(), "Cheques")
                df_ch = df_ch[df_ch['email']==cliente_mail].copy() if not df_ch.empty and 'email' in df_ch.columns else pd.DataFrame()
            except Exception:
                df_ch = pd.DataFrame()

            if df_ch.empty:
                st.info("Sin cheques registrados aún.")
            else:
                df_ch['monto']      = pd.to_numeric(df_ch['monto'], errors='coerce').fillna(0)
                df_ch['fecha_venc_dt'] = pd.to_datetime(df_ch['fecha_venc'], dayfirst=True, errors='coerce')
                hoy = date.today()

                # ── ALERTAS: cheques próximos a vencer ──
                df_ch_pend = df_ch[df_ch['estado'] != 'Depositado'] if 'estado' in df_ch.columns else df_ch
                hoy_ts2 = pd.Timestamp(hoy)
                limite_ts = hoy_ts2 + pd.Timedelta(days=int(dias_alerta))
                df_prox = df_ch_pend[
                    df_ch_pend['fecha_venc_dt'].notna() &
                    (df_ch_pend['fecha_venc_dt'] >= hoy_ts2) &
                    (df_ch_pend['fecha_venc_dt'] <= limite_ts)
                ] if not df_ch_pend.empty else pd.DataFrame()

                df_venc = df_ch_pend[
                    df_ch_pend['fecha_venc_dt'].notna() &
                    (df_ch_pend['fecha_venc_dt'] < hoy_ts2)
                ] if not df_ch_pend.empty else pd.DataFrame()

                if not df_venc.empty:
                    st.error(f"🚨 {len(df_venc)} cheque(s) VENCIDO(S) sin depositar!")
                    for _,r in df_venc.iterrows():
                        st.markdown(f"**N° {r['numero']}** — {r['banco']} — {fmt_ar(r['monto'])} — Venció: {r['fecha_venc']}")

                if not df_prox.empty:
                    dias_str = f"{dias_alerta} días"
                    st.warning(f"⚠️ {len(df_prox)} cheque(s) próximo(s) a vencer en {dias_str}!")
                    for _,r in df_prox.iterrows():
                        dias_rest = (r['fecha_venc_dt'] - hoy_ts2).days
                        st.markdown(
                            f"<div style='background:#FEF9E7; border-left:4px solid #F39C12; "
                            f"border-radius:8px; padding:10px 14px; margin:4px 0;'>"
                            f"🔔 <strong>N° {r['numero']}</strong> — {r['banco']} — "
                            f"<strong>{fmt_ar(r['monto'])}</strong> — "
                            f"Vence: {r['fecha_venc']} (<strong style='color:#E74C3C;'>faltan {dias_rest} días</strong>)"
                            f"</div>",
                            unsafe_allow_html=True
                        )

                st.divider()

                # ── Tabla de cheques ──
                col_ch1, col_ch2 = st.columns(2)
                with col_ch1:
                    st.markdown("##### 📥 Cheques Recibidos (para depositar)")
                    df_rec = df_ch[df_ch['tipo']=='Cheque Recibido'] if 'tipo' in df_ch.columns else df_ch
                    if df_rec.empty:
                        st.info("Sin cheques recibidos")
                    else:
                        for _,r in df_rec.sort_values('fecha_venc_dt').iterrows():
                            estado_color = '#27AE60' if str(r.get('estado','')) == 'Depositado' else '#E74C3C'
                            st.markdown(
                                f"<div style='background:white; border-radius:10px; padding:10px 14px; "
                                f"margin:4px 0; border-left:4px solid {estado_color}; "
                                f"box-shadow:0 1px 4px rgba(0,0,0,0.08);'>"
                                f"<strong>N° {r['numero']}</strong> — {r.get('banco','')} — {fmt_ar(r['monto'])}<br>"
                                f"<small style='color:#7F8C8D;'>Vence: {r['fecha_venc']} | "
                                f"Librador: {r.get('librador','')} | Estado: {r.get('estado','Pendiente')}</small>"
                                f"</div>",
                                unsafe_allow_html=True
                            )
                            # Botón marcar como depositado
                            if str(r.get('estado','')) != 'Depositado':
                                if st.button(f"✅ Marcar depositado N°{r['numero']}", key=f"dep_{r['id']}"):
                                    try:
                                        df_ch_full = read_ws(open_sheet(), "Cheques")
                                        df_ch_full.loc[df_ch_full['id'].astype(str)==str(r['id']), 'estado'] = 'Depositado'
                                        write_ws("Cheques", df_ch_full)
                                        st.cache_data.clear(); st.session_state.pop('data_cache', None); st.rerun()
                                    except Exception as e:
                                        st.error(f"Error: {e}")

                with col_ch2:
                    st.markdown("##### 📤 Cheques Emitidos (pagos realizados)")
                    df_emit = df_ch[df_ch['tipo']=='Cheque Emitido'] if 'tipo' in df_ch.columns else pd.DataFrame()
                    if df_emit.empty:
                        st.info("Sin cheques emitidos")
                    else:
                        for _,r in df_emit.sort_values('fecha_venc_dt').iterrows():
                            st.markdown(
                                f"<div style='background:white; border-radius:10px; padding:10px 14px; "
                                f"margin:4px 0; border-left:4px solid #3498DB; "
                                f"box-shadow:0 1px 4px rgba(0,0,0,0.08);'>"
                                f"<strong>N° {r['numero']}</strong> — {r.get('banco','')} — {fmt_ar(r['monto'])}<br>"
                                f"<small style='color:#7F8C8D;'>Vence: {r['fecha_venc']} | "
                                f"A favor de: {r.get('librador','')} | Estado: {r.get('estado','Pendiente')}</small>"
                                f"</div>",
                                unsafe_allow_html=True
                            )

                # Métricas resumen
                st.divider()
                mc1,mc2,mc3,mc4 = st.columns(4)
                df_rec_pend = df_ch[(df_ch.get('tipo','')=='Cheque Recibido') & (df_ch.get('estado','Pendiente')!='Depositado')] if not df_ch.empty else pd.DataFrame()
                mc1.metric("📥 Cheques recibidos", str(len(df_ch[df_ch.get('tipo','')=='Cheque Recibido'])) if 'tipo' in df_ch.columns else "0")
                mc2.metric("💰 Total a depositar", fmt_ar(df_rec_pend['monto'].sum()) if not df_rec_pend.empty else "$ 0")
                mc3.metric("📤 Cheques emitidos",  str(len(df_ch[df_ch.get('tipo','')=='Cheque Emitido'])) if 'tipo' in df_ch.columns else "0")
                mc4.metric("⚠️ Próximos a vencer", str(len(df_prox)))

    # ══════════════════════════════════════════
    # ══════════════════════════════════════════
    # INVERSIONES
    # ══════════════════════════════════════════
    elif menu == "📈 Inversiones":
        st.markdown(f'<div class="page-header"><div><h2>📈 Inversiones</h2><span>{sel_nombre}</span></div></div>', unsafe_allow_html=True)
        mis_inv = df_inv_raw[df_inv_raw['email']==cliente_mail].copy() if not df_inv_raw.empty else pd.DataFrame()
        if not mis_inv.empty:
            if 'monto' in mis_inv.columns:
                mis_inv['monto'] = pd.to_numeric(mis_inv['monto'],errors='coerce').fillna(0)
            if 'rentabilidad' not in mis_inv.columns: mis_inv['rentabilidad'] = 0
            if 'tipo_registro' not in mis_inv.columns: mis_inv['tipo_registro'] = 'recomendacion_admin'
        regs = mis_inv[mis_inv['tipo_registro']=='registro_cliente'] if not mis_inv.empty else pd.DataFrame()
        recs = mis_inv[mis_inv['tipo_registro']=='recomendacion_admin'] if not mis_inv.empty else pd.DataFrame()
        tot_inv = regs['monto'].sum() if not regs.empty and 'monto' in regs.columns else 0
        if tot_inv > 0:
            mi1,mi2,mi3,_ = st.columns(4)
            mi1.metric("📦 TOTAL INVERTIDO", fmt_ar(tot_inv))
            if not regs.empty and 'rentabilidad' in regs.columns:
                _rv = pd.to_numeric(regs['rentabilidad'],errors='coerce').dropna()
                mi2.metric("📈 RENT. PROM.", f"{_rv.mean():.1f}%" if not _rv.empty else "N/A")
            mi3.metric("🔢 POSICIONES", str(len(regs)))
        tabs_inv = st.tabs(["📝 Mis Registros","💡 Recomendaciones"] if not es_admin
                           else ["📝 Registros Cliente","💡 Enviar Recomendación","📊 Resumen Admin"])
        with tabs_inv[0]:
            if not es_admin:
                st.markdown("#### ➕ Registrar inversión")
                with st.form("reg_inv"):
                    ri1,ri2,ri3 = st.columns(3)
                    ins_r  = ri1.selectbox("📊 Instrumento",["Bonos","Acciones","CEDEARs","FCI","Dólar MEP","Cripto","Plazo Fijo","Otro"])
                    mon_r  = ri2.number_input("💵 Monto ($)",min_value=0.0,step=1000.0)
                    rent_r = ri3.number_input("📈 Rentabilidad (%)",min_value=-100.0,max_value=1000.0,step=0.1)
                    nota_r = st.text_input("📝 Nombre del fondo / detalle")
                    if st.form_submit_button("💾 REGISTRAR",use_container_width=True):
                        ni = pd.DataFrame([{"email":cliente_mail,"fecha":date.today().strftime('%d/%m/%Y'),
                            "instrumento":ins_r,"monto":mon_r,"rentabilidad":rent_r,
                            "mensaje":nota_r,"tipo_registro":"registro_cliente"}])
                        write_ws("Inversiones", pd.concat([df_inv_raw,ni],ignore_index=True))
                        st.success("✅ Inversión registrada"); st.cache_data.clear(); st.rerun()
            if regs.empty:
                st.info("📭 Sin inversiones registradas.")
            else:
                for _,r in regs.sort_index(ascending=False).iterrows():
                    ci1,ci2,ci3 = st.columns([2,1,1])
                    ci1.markdown(f'<div class="inv-card"><strong>{r["instrumento"]}</strong><br><small style="color:#95A5A6;">{r["fecha"]}</small></div>', unsafe_allow_html=True)
                    ci2.metric("Monto",fmt_ar(r['monto']))
                    rv = float(r.get('rentabilidad',0)) if pd.notna(r.get('rentabilidad',0)) else 0
                    ci3.metric("Rentabilidad",f"{rv:.1f}%",delta=f"{rv:.1f}%" if rv!=0 else None)
        with tabs_inv[1]:
            if es_admin:
                st.markdown("#### 📤 Enviar recomendación")
                with st.form("rec_admin"):
                    ra1,ra2 = st.columns(2)
                    ins_a = ra1.selectbox("📊 Instrumento",["Dólar MEP","FCI","CEDEAR","Plazo Fijo","Bonos","Cripto","Acciones","Otro"])
                    mon_a = ra2.number_input("💵 Monto Sugerido ($)",min_value=0.0,step=1000.0)
                    msg_a = st.text_area("💬 Análisis / Recomendación")
                    if st.form_submit_button("📤 ENVIAR",use_container_width=True):
                        ni = pd.DataFrame([{"email":cliente_mail,"fecha":date.today().strftime('%d/%m/%Y'),
                            "instrumento":ins_a,"monto":mon_a,"rentabilidad":0,
                            "mensaje":msg_a,"tipo_registro":"recomendacion_admin"}])
                        write_ws("Inversiones", pd.concat([df_inv_raw,ni],ignore_index=True))
                        st.success("✅ Recomendación enviada"); st.cache_data.clear()
            else:
                if recs.empty:
                    st.info("📭 Tu contador/a aún no te envió recomendaciones.")
                else:
                    for _,r in recs.sort_index(ascending=False).iterrows():
                        with st.expander(f"📊 {r['instrumento']} — {r['fecha']}"):
                            rc1,rc2 = st.columns([1,2])
                            rc1.metric("Monto sugerido",fmt_ar(r['monto']))
                            rc2.markdown(f"**Análisis:**\n\n{r.get('mensaje','')}")
        if es_admin and len(tabs_inv) > 2:
            with tabs_inv[2]:
                if regs.empty:
                    st.info("El cliente no tiene inversiones registradas.")
                else:
                    ra1,ra2,ra3 = st.columns(3)
                    ra1.metric("Total Invertido", fmt_ar(pd.to_numeric(regs['monto'],errors='coerce').fillna(0).sum()))
                    ra2.metric("Posiciones", str(len(regs)))
                    ra3.metric("Recomendaciones", str(len(recs)))

    # ══════════════════════════════════════════
    # PERFIL
    # ══════════════════════════════════════════
    elif menu == "⚙️ Perfil":
        st.markdown(f'<div class="page-header"><div><h2>⚙️ Perfil</h2><span>{user["nombre"]} · {user["email"]}</span></div></div>', unsafe_allow_html=True)

        # Admin tiene tab extra para la frase semanal
        if es_admin:
            tab_pf1, tab_pf2, tab_pf3 = st.tabs(["🔒 Contraseña", "📱 Mensaje WhatsApp", "✨ Frase Semanal"])
        else:
            tab_pf1, tab_pf2 = st.tabs(["🔒 Contraseña", "📱 Mensaje WhatsApp"])

        with tab_pf1:
            col_pf,_ = st.columns([1,2])
            with col_pf:
                st.markdown("#### 🔒 Cambiar contraseña")
                with st.form("p"):
                    n1 = st.text_input("Nueva contraseña", type="password")
                    n2 = st.text_input("Repetir contraseña", type="password")
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.form_submit_button("CAMBIAR CONTRASEÑA", use_container_width=True):
                        if n1!=n2: st.error("No coinciden")
                        elif len(n1)<=3: st.warning("Minimo 4 caracteres")
                        else:
                            df_u = data['users'].copy()
                            df_u.loc[df_u['email']==user['email'],'password'] = n1
                            write_ws("Users", df_u)
                            st.success("Contraseña actualizada. Reingresa.")
                            st.session_state.clear(); st.cache_data.clear(); st.rerun()

        with tab_pf2:
            st.markdown("#### 📱 Personalizar mensaje de recordatorio WhatsApp")
            tpl_default = "Hola {nombre}! Te recordamos que tenes un pago pendiente de {monto}. Muchas gracias!"
            tpl_key     = f"wa_template_{user['email']}"
            tpl_actual  = st.session_state.get(tpl_key, st.session_state.get('wa_template', tpl_default))
            col_wa1, col_wa2 = st.columns([3,2])
            with col_wa1:
                nuevo_tpl = st.text_area("Texto del mensaje:", value=tpl_actual, height=130,
                    help="Usa {nombre} y {monto} como variables automaticas.")
                if st.button("Guardar mensaje de forma permanente", use_container_width=True):
                    st.session_state[tpl_key] = nuevo_tpl
                    st.session_state['wa_template'] = nuevo_tpl
                    df_u = data['users'].copy()
                    if 'wa_template' not in df_u.columns: df_u['wa_template'] = ''
                    df_u.loc[df_u['email'].astype(str).str.lower().str.strip()==user['email'],'wa_template'] = nuevo_tpl
                    write_ws("Users", df_u)
                    st.cache_data.clear()
                    st.success("Mensaje guardado!")
                if st.button("Restaurar mensaje por defecto", use_container_width=True):
                    st.session_state[tpl_key] = tpl_default
                    st.session_state['wa_template'] = tpl_default
                    st.rerun()
            with col_wa2:
                st.markdown("##### Vista previa")
                preview = nuevo_tpl.replace("{nombre}", "Juan Perez").replace("{monto}", "$ 15.000")
                st.markdown(
                    f"<div style='background:#DCF8C6; border-radius:12px 12px 3px 12px;"
                    f"padding:12px 16px; font-size:0.9rem; color:#1A252F;"
                    f"box-shadow:0 2px 6px rgba(0,0,0,0.12); max-width:280px; margin-top:8px;'>"
                    + preview +
                    "</div><div style='color:#95A5A6; font-size:0.75rem; margin-top:4px;"
                    "text-align:right; max-width:280px;'>Enviado</div>",
                    unsafe_allow_html=True)

        # ── TAB FRASE SEMANAL (solo admin) ──
        if es_admin:
            with tab_pf3:
                st.markdown("#### ✨ Frase semanal para los clientes")
                st.markdown("""
                <div style='background:#F0F7FF; border-left:4px solid #1B4F8A;
                            border-radius:8px; padding:10px 14px; margin-bottom:12px; font-size:0.85rem;'>
                Al guardar, la frase aparece como <strong>popup al ingresar</strong> y queda
                visible en la parte superior del dashboard de cada cliente durante toda la semana.
                </div>
                """, unsafe_allow_html=True)

                # Leer frase actual del Config sheet (columna 'frase_semanal', fila 1)
                frase_actual = st.session_state.get('frase_semanal_admin', '')
                if not frase_actual:
                    try:
                        if 'frase_semanal' in df_config.columns:
                            _f = df_config['frase_semanal'].dropna()
                            frase_actual = str(_f.iloc[0]) if not _f.empty else ''
                    except Exception:
                        frase_actual = ''

                col_fr1, col_fr2 = st.columns([3,2])
                with col_fr1:
                    nueva_frase = st.text_area(
                        "✏️ Escribí la frase de esta semana:",
                        value=frase_actual, height=120,
                        placeholder="Ej: El éxito financiero no es un destino, es un hábito diario. ¡Vamos juntos! 💪"
                    )
                    autor_frase = st.text_input("Autor / Firma (opcional):",
                        placeholder="Ej: — Tu contadora, Jessica")

                    if st.button("💾 Publicar frase de la semana", use_container_width=True, key="btn_pub_frase"):
                        frase_completa = nueva_frase
                        if autor_frase:
                            frase_completa = nueva_frase + f"\n\n— {autor_frase}"
                        # Guardar en Config sheet
                        try:
                            df_cfg_upd = df_config.copy()
                            if 'frase_semanal' not in df_cfg_upd.columns:
                                df_cfg_upd['frase_semanal'] = ''
                            df_cfg_upd.loc[0, 'frase_semanal'] = frase_completa
                            write_ws("Config", df_cfg_upd)
                            st.session_state['frase_semanal_admin'] = frase_completa
                            st.cache_data.clear()
                            st.session_state.pop('data_cache', None)
                            st.success("✅ Frase publicada! Los clientes la verán al ingresar.")
                        except Exception as _e:
                            st.error(f"Error al guardar: {_e}")

                    if st.button("🗑️ Quitar frase", use_container_width=True, key="btn_del_frase"):
                        try:
                            df_cfg_upd = df_config.copy()
                            if 'frase_semanal' in df_cfg_upd.columns:
                                df_cfg_upd['frase_semanal'] = ''
                            write_ws("Config", df_cfg_upd)
                            st.session_state['frase_semanal_admin'] = ''
                            st.cache_data.clear()
                            st.session_state.pop('data_cache', None)
                            st.success("Frase quitada.")
                        except Exception as _e:
                            st.error(f"Error: {_e}")

                with col_fr2:
                    st.markdown("##### 👁️ Vista previa del popup")
                    _prev_frase = nueva_frase if nueva_frase else "Tu frase aparecerá aquí..."
                    _prev_autor = f"\n\n— {autor_frase}" if autor_frase else ""
                    st.markdown(f"""
                    <div style='background:linear-gradient(135deg,#1B4F8A,#2980B9);
                                border-radius:16px; padding:24px 20px; text-align:center;
                                box-shadow:0 8px 24px rgba(27,79,138,0.35); margin-top:8px;'>
                        <div style='font-size:1.8rem; margin-bottom:8px;'>✨</div>
                        <p style='color:white; font-size:0.95rem; font-style:italic;
                                  line-height:1.6; margin:0;'>{_prev_frase}</p>
                        <p style='color:#BDC3C7; font-size:0.8rem; margin:12px 0 0 0;'>{_prev_autor.replace(chr(10),'')}</p>
                    </div>
                    """, unsafe_allow_html=True)
